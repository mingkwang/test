import os
import gc
import os.path
import sys
import threading
import time
import tkinter.messagebox
import tkinter.scrolledtext
from ctypes import *
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import pandas as pd
import serial
import serial.tools.list_ports
import re
import openpyxl
class Log(object):
    def __init__(self):
        self.file = './log'
        if not os.path.exists(self.file):
            os.makedirs('./log')
        self.filename = self.file + '/log-' + time.strftime('%Y-%m-%d_%H-%M-%S') + '.txt'
        self.log = open(self.filename, 'a+', encoding='UTF-8')

    def write(self, message, scr, mode=1):
        self.log = open(self.filename, 'a', encoding='UTF-8')
        self.log.write(time.strftime('%Y-%m-%d_%H-%M-%S') + ': ' + message + '\n')
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ': ' + message)
        if mode == 0:
            return 0
        scr.insert(END, time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '→' + message + '\n')
        scr.see(END)
        scr.update()
        self.close()

    def close(self):
        self.log.close()
class callback:
    def __init__(self, dll):
        self.dll = dll
        return

    def Function(self, callback, p):
        return self.dll.ccrfidDevSdkStartWork(callback, p)

    def write_AT(self, callback):
        pass
class SimpStruct(Structure):
    _fields_ = [("cls", c_ulong),  # 远程设备class
                ("dev_hdl", c_ulong),  # 远程设备的device handle
                ("name", c_char * 32),  # 远程设备名字
                ("BdAddr", c_ubyte * 6),  # 远程设备地址
                ("hConn", c_ulong),  # 远程设备connection handle
                ("placement", c_char * 10)]
class ATTestList(Structure):
    pass
ATTestList._fields_ = [("ATCmdSingleCmdStr", c_char_p),  # 定义一维数组,字符串数组
                       ("ATCmdSinglelCmdLength", c_uint16),
                       ("ATCmdSingleCmdResultStr", c_char_p),
                       ("ATCmdSingleCmdFeedbackStr", c_char_p),
                       ("ATCmdCheckoutResult", c_bool),
                       ("Next", POINTER(ATTestList))]
class LmpInfo(Structure):
    _fields_ = [("lmp_feature", c_ubyte * 8),
                ("manuf_name", c_ushort),
                ("lmp_subversion", c_ushort),
                ("lmp_version", c_ubyte),
                ("hci_version", c_ubyte),
                ("hci_revision", c_ushort),
                ("country_code", c_ubyte)]
class LocalInfo(Structure):
    _fields_ = [("name", c_char * 32),
                ("discover_mode", c_ushort),
                ("pin_code", c_char * 32),
                ("device_class", c_ulong),
                ("bd_addr", c_ubyte * 6),
                ("lmp_info", LmpInfo),
                ("app_param", c_ulong)]
class BtSdkUUIDStru(Structure):
    _fields_ = [("Data1", c_ulong),
                ("Data2", c_ushort),
                ("Data3", c_ushort),
                ("Data4", c_char * 8)]
class BtsdkGATTUUIDStru(Structure):
    _fields_ = [("IsShortUuid", c_long),
                ("ShortUuid", c_ushort),
                ("LongUuid", BtSdkUUIDStru)]
class BtsdkGATTServiceStru(Structure):
    _fields_ = [("ServiceUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort)]
class BtSdkGATTService(Structure):
    _fields_ = [("num", c_ushort),
                ("service", BtsdkGATTServiceStru * 10)]
class BtsdkGATTCharacteristicStru(Structure):
    _fields_ = [("ServiceHandle", c_ushort),
                ("CharacteristicUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort),
                ("CharacteristicValueHandle", c_ushort),
                ("IsBroadcastable", c_ulong),
                ("IsReadable", c_ulong),
                ("IsWritable", c_ulong),
                ("IsWritableWithoutRespon", c_ulong),
                ("IsSignedWritable", c_ulong),
                ("IsNotifiable", c_ulong),
                ("IsIndicatable", c_ulong),
                ("HasExtendedProperties", c_ulong)]
class BtSdkGATTCharacteristic(Structure):
    _fields_ = [("num", c_ushort),
                ("character", BtsdkGATTCharacteristicStru * 10)]
class BtsdkGATTDescriptorStru(Structure):
    _fields_ = [("ServiceHandle", c_ushort),
                ("CharacteristicHandle", c_ushort),
                ("DescriptorType", c_ulong),
                ("DescriptorUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort)]
class BtSdkGATTDescriptor(Structure):
    _fields_ = [("num", c_ushort),
                ("des", BtsdkGATTDescriptorStru * 10)]
class CEP(Structure):
    _fields_ = [("IsReliableWriteEnabled", c_ulong),
                ("IsAuxiliariesWritable", c_ulong)]
class CCC(Structure):
    _fields_ = [("IsSubscribeToNotification", c_ulong),
                ("IsSubscribeToIndication", c_ulong)]
class SCC(Structure):
    _fields_ = [("IsBoardcast", c_ulong)]
class CF(Structure):
    _fields_ = [("Format", c_char),
                ("Exponent", c_char),
                ("Uint", BtsdkGATTUUIDStru),
                ("NameSpace", c_char),
                ("Descriptor", BtsdkGATTUUIDStru)]
class U(Union):
    _fields_ = [("CharacterExtendedProperties", CEP),
                ("ClientCharacteristicConfiguration", CCC),
                ("ServerCharacteristicConfiguration", SCC),
                ("CharacteristicFormat", CF)]
class BtsdkGATTDscriptorValueStru(Structure):
    _fields_ = [("DescriptorType", c_ulong),
                ("DescriptorUuid", BtsdkGATTUUIDStru),
                ("u", U),
                ("DataSize", c_ulong),
                ("Data", c_char * 100)]
class BtsdsGATTCharacteristicValueStru(Structure):
    _fields_ = [("DataSize", c_ulong),
                ("Data", c_ubyte * 100)]
class ServerStruct(Structure):
    _fields_= [ ("data",      c_char*500),     #接收的数据
                ("data_size", c_ushort  )]

class Auto_Test(object):

    def __init__(self):
        self.delete_ini_file()
        self.root = Tk()
        self.times = 0
        self.root.geometry('800x690+400+50')
        self.root.resizable(0, 0)
        self.root.title('Automatic TestReport generation tool')
        self.ini = '初始化条件'
        self.process = '测试过程'
        self.result = '测试结果'
        self.expecte = '期望结果'
        self.question = '问题描述'
        # self.sheet_name = 'SPP'
        self.com_list = []
        self.conn_flag = 0
        self.report_path = 'test_Report.xlsx'
        self.result_data = []  # 把data转换为二维数组(没有必要)
        self.type1 = ['用例ID', '类型', '级别', '目的', '初始化条件', '测试过程', '期望结果', '测试结果', '问题描述', '概率',
                     '测试手机']  # 可以先用 xlrd 读第一行的值
        self.log = Log()
        self.search_flag = 0
        self.flow_control_flag = 0
        self.dev_hdl = 0
        self.hConn = 0
        self.conns = []
        self.at_res = ''
        self.isconnflag = 0
        self.isPDU = 0
        self.devices_id = 0  # remote_index
        self.conn_num_spp = 0
        self.conn_num_gatt = 0
        self.hConn_info_gatt = []   # 整合 gatt spp device的列表为一个
        self.device_info = []  # 多连接 存放dangle的信息（device_id, gatt_dev_hdl, gatt_bd, gatt_conn,spp_dev_hdl, spp_conn_hdl, spp_conn, device_bd）
        self.hConn_info = []  # 存储remote连接的搜索句柄，连接句柄[{id:0,dev_hdl:搜索句柄,conn_hdl:连接句柄.....},
        self.sdk_dll = CDLL("sdksample.dll")
        self.init()
        self.pDll_su = CDLL("ATTestLib.dll")

    def init(self):
        self.sdk_dll.BluetDisConnectRemoteDevice.argtype = [c_ulong]
        self.sdk_dll.BluetDisConnectRemoteDevice.retypes = c_ushort
        self.sdk_dll.BluetFuncInterfaceInit()
        self.sdk_dll.GATTSeriveInit()
        self.sdk_dll.Get_Version.retypes = c_char_p
        version = self.sdk_dll.Get_Version()
        print('当前版本：' + str(version))

    def SearchCompleteCback(self, dev_hdl, bd_addr):
        self.search_flag = 1
        self.dev_hdl = dev_hdl

    def input_pincode(self, dev_hdl, pin):
        self.sdk_dll.SetRemoteDevicePinCode.argtypes = [c_char_p, c_uint16]
        self.sdk_dll.SetRemoteDevicePinCode(bytes(pin, 'utf-8'), c_uint16(len(pin)))
        self.sdk_dll.BluetPairDevice(dev_hdl)

    def strtohex(self, a):
        '''
        字符串转换16进制
        :param a:
        :return:
        '''
        sum1 = 0
        length = len(a) - 1
        for i in range(len(a)):
            if '0' <= a[i] <= '9':
                sum1 += int(a[i]) * (16 ** (length - i))
            if a[i] == 'A' or a[i] == 'a':
                sum1 += 10 * (16 ** (length - i))
            if a[i] == 'B' or a[i] == 'b':
                sum1 += 11 * (16 ** (length - i))
            if a[i] == 'C' or a[i] == 'c':
                sum1 += 12 * (16 ** (length - i))
            if a[i] == 'D' or a[i] == 'd':
                sum1 += 13 * (16 ** (length - i))
            if a[i] == 'E' or a[i] == 'e':
                sum1 += 14 * (16 ** (length - i))
            if a[i] == 'F' or a[i] == 'f':
                sum1 += 15 * (16 ** (length - i))

        return hex(sum1)

    def str_to_addr(self, addr):
        addr_back = []
        if 14 >= len(addr) >= 12:
            for i in range(12, 0, -2):
                addr_back.append(int(self.strtohex(addr[i - 2:i]), 16))
            return addr_back
        else:
            print('连接设备地址错误！')
            return 0

    def read_excel(self):
        '''
        读取excel里面的数据
        :return:
        '''
        self.data = None
        self.data = pd.read_excel(self.path, sheet_name=self.sheet_name)
        self.type = self.data.columns.values.tolist()            # 得到第一行的值
        length = len(self.data[self.type[0]])
        for i in range(length):
            temp = []
            for j in range(len(self.type)):
                data_temp = self.data[self.type[j]][i]
                temp.append(data_temp)
            self.result_data.append(temp)

    def make_data(self, index, data):
        result = []
        for i in range(len(data)):
            temp_s = self.result_data[i][index]
            temp_d = []
            if '\CR' in temp_s:
                temp_s = temp_s.split('\CR')
                for reg in temp_s:
                    reg = reg.upper()  # 全转换为大写字母
                    if 'SOC' in reg:
                        reg = reg.split('SOC')  # 拆分AT指令
                        reg = 'soc' + reg[1]
                        temp_d.append(reg)
                    elif 'soc' in reg:
                        reg = reg.split('soc')
                        reg = 'soc' + reg[1]
                        temp_d.append(reg)
                    elif 'REMOTE' in reg:
                        reg = reg.split('REMOTE')
                        reg = 'remote' + reg[1]
                        temp_d.append(reg)
                    elif 'remote' in reg:
                        reg = reg.split('remote')
                        reg = 'remote' + reg[1]
                        temp_d.append(reg)
                    elif 'LOCAL' in reg:
                        reg = reg.split('LOCAL')
                        reg = 'local' + reg[1]
                        temp_d.append(reg)
                    elif 'local' in reg:
                        reg = reg.split('local')
                        reg = 'local' + reg[1]
                        temp_d.append(reg)
            result.append(temp_d)

        return result

    def del_data(self):
        '''
        处理测试用例的表格数据，得到AT指令
        :return:
        '''
        self.read_excel()
        index = self.type.index(self.ini)
        index1 = self.type.index(self.process)
        result_order = self.make_data(index, self.result_data)
        result_order1 = self.make_data(index1, self.result_data)
        return result_order, result_order1

    def get_port(self):
        '''
        检索串口
        :return
        '''
        self.com_list.clear()
        plist = list(serial.tools.list_ports.comports())
        for por in range(0, len(plist)):
            if str(plist[por]).find('COM') >= 0:
                com = str(plist[por])[0:4]
                self.com_list.append(com)
        if len(self.com_list) == 0:
            self.log.write('没有找到对应模块的串口', self.scr)

    def init_spp_service(self, count):
        '''
        初始化spp服务，
        :param count:
        :return:
        '''
        self.sdk_dll.BluetRegisterSPPService.argtype = [c_ubyte]
        self.sdk_dll.BluetRegisterSPPService(count)

    def Disconnection(self, conn, del_hdl=1):
        '''
        断开连接
        :param del_hdl:  是否删除已连接的信息
        :param conn:
        :return:
        '''
        self.log.write('断开SPP连接', self.scr)
        flag = self.sdk_dll.BluetDisConnectRemoteDevice(conn)
        if flag != 1:
            return 0
        self.hConn_info[self.devices_id]['spp'] = 0
        self.conn_num_spp -= 1
        self.log.write('当前SPP连接数量：'+str(self.conn_num_spp), self.scr, mode=1)
        return 1

    def Connection(self, braddr, brclass=0, pin='1234', mode=13):
        '''
        连接手机
        :return:
        '''
        # 定义回调函数
        CALLBACK = CFUNCTYPE(None, c_ulong, POINTER(c_ubyte))
        callBackFunc = CALLBACK(self.SearchCompleteCback)
        self.sdk_dll.SearchCompleCallback(callBackFunc)
        self.sdk_dll.BluetSearchRemoteDevice.argtype = [c_ulong, c_char_p]
        self.sdk_dll.BluetSearchRemoteDevice.restype = c_ulong
        brclass = c_ulong(int(brclass))
        addr = self.str_to_addr(braddr)

        self.sdk_dll.BluetSetLocalMode.argtypes = [c_ushort]
        mode1 = c_ushort(mode)
        self.sdk_dll.BluetSetLocalMode(mode1)  # 设置手机蓝牙发现模式

        self.sdk_dll.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
        pin_code = c_char_p(bytes(pin, 'utf-8'))
        pin_code_len = len(pin)
        self.sdk_dll.BluetSetFixedPinCode(pin_code, pin_code_len)

        if addr == 0:
            print('连接设备的蓝牙地址错误！')
            self.log.write('连接设备的蓝牙地址错误', self.scr)
            return 0
        bdaddr = (c_char * 6)(*addr)
        Search_flag = self.sdk_dll.BluetSearchRemoteDevice(brclass, bdaddr)
        print('模块搜索中！')
        time_flag = 0
        while True:
            time.sleep(2)
            time_flag += 2
            if time_flag >= 20:
                return 0
            if self.search_flag == 1:
                self.search_flag = 0
                print('已经搜索到经典蓝牙设备')
                self.log.write('已经搜索到经典蓝牙设备', self.scr)
                self.sdk_dll.BluetStopDiscoverDevice()
                self.sdk_dll.BluetConnectRemoteDevice.argtype = [c_ulong]
                self.sdk_dll.BluetConnectRemoteDevice.restype = c_ulong
                print('设备句柄：' + str(self.dev_hdl))
                self.input_pincode(self.dev_hdl,pin)  # 输入pincode
                conn_hdl = self.sdk_dll.BluetConnectRemoteDevice(self.dev_hdl)
                self.hConn = conn_hdl
                print('连接句柄')
                print(conn_hdl)
                if conn_hdl == 13 or conn_hdl == 0:
                    self.log.write('模块SPP连接失败', self.scr)
                    return 0
                else:
                    if self.devices_id+1 > len(self.hConn_info):
                        dev_dic = {}
                        dev_dic['id'] = self.devices_id
                        dev_dic['dev_hdl'] = self.dev_hdl
                        dev_dic['conn_hdl'] = conn_hdl
                        dev_dic['bd'] = braddr
                        dev_dic['spp'] = 1
                        self.hConn_info.append(dev_dic)
                    else:
                        self.hConn_info[self.devices_id]['dev_hdl'] = self.dev_hdl
                        self.hConn_info[self.devices_id]['conn_hdl'] = conn_hdl
                        self.hConn_info[self.devices_id]['bd'] = braddr
                        self.hConn_info[self.devices_id]['spp'] = 1
                    self.conn_num_spp += 1
                self.conns.append(self.hConn)
                return 1

    def Get_Gatt_Services(self, dev_handle):
        self.sdk_dll.GATT_GetServices.argtypes = [c_ulong]
        self.sdk_dll.GATT_GetServices.restype = POINTER(BtSdkGATTService)

        self.sdk_dll.ATT_GETUUIDType.argtypes = [c_ushort]
        self.sdk_dll.ATT_GETUUIDType.restype = c_char_p

        self.sdk_dll.GATT_GetCharacteristics.argtypes = [c_ubyte]  # 得到特征 get characteristics
        self.sdk_dll.GATT_GetCharacteristics.restype = POINTER(BtSdkGATTCharacteristic)

        self.sdk_dll.GATT_GetCharacteristicsValue.argtypes = [c_ubyte]  # 得到特征值 get characteristics_value
        self.sdk_dll.GATT_GetCharacteristicsValue.restype = POINTER(BtsdsGATTCharacteristicValueStru)

        self.sdk_dll.GATT_SetCharacteristicsValue.argtypes = [c_char, c_char]  # 设置特征值 set characteristics_value
        self.sdk_dll.GATT_SetCharacteristicsValue.retypes = c_long

        self.sdk_dll.GATT_GetDescriptors.argtypes = [c_ubyte]  # 得到描述 get Descriptors
        self.sdk_dll.GATT_GetDescriptors.restype = POINTER(BtSdkGATTDescriptor)

        self.sdk_dll.GATT_GetDescriptorValue.argtypes = [c_ubyte]  # 得到描述值 get Descriptors_value
        self.sdk_dll.GATT_GetDescriptorValue.restype = POINTER(BtsdkGATTDscriptorValueStru)

        self.sdk_dll.GATT_SetDescriptorValue.argtypes = [c_ubyte, c_ubyte]  # 设置描述值 set Descriptors_value

        try:
            gatt_service = self.sdk_dll.GATT_GetServices(c_ulong(dev_handle))  # 得到服务 get services
            if gatt_service is None:
                print('没有发现服务')
                self.log.write('没有发现服务', self.scr)
                return '没有发现服务'
            else:
                for i in range(int(gatt_service.contents.num)):
                    uuid = gatt_service.contents.service[i].ServiceUuid.ShortUuid
                    src = self.sdk_dll.ATT_GETUUIDType(uuid)
                    print("service id:%d  uuid = %d %s handle = 0x%02x" % (
                        i + 1, uuid, src.decode(), gatt_service.contents.service[i].AttributeHandle))
                    self.log.write("service id:%d  uuid = %d %s handle = 0x%02x" % (
                        i + 1, uuid, src.decode(), gatt_service.contents.service[i].AttributeHandle), self.scr)
                    gatt_character = self.sdk_dll.GATT_GetCharacteristics(c_ubyte(i + 1))
                    if gatt_character.contents is not None:
                        for j in range(gatt_character.contents.num):
                            uuid_c = gatt_character.contents.character[j].CharacteristicUuid.ShortUuid
                            src_c = self.sdk_dll.ATT_GETUUIDType(uuid_c)
                            self.log.write((" character id:%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
                                j + 1, gatt_character.contents.character[j].AttributeHandle,
                                gatt_character.contents.character[j].CharacteristicValueHandle, src_c.decode())), self.scr)
                            print(" character id:%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
                                j + 1, gatt_character.contents.character[j].AttributeHandle,
                                gatt_character.contents.character[j].CharacteristicValueHandle, src_c.decode()))

                            self.log.write("  [B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
                                gatt_character.contents.character[j].IsBroadcastable,
                                gatt_character.contents.character[j].IsReadable,
                                gatt_character.contents.character[j].IsWritable,
                                gatt_character.contents.character[j].IsSignedWritable,
                                gatt_character.contents.character[j].IsWritableWithoutRespon,
                                gatt_character.contents.character[j].IsIndicatable,
                                gatt_character.contents.character[j].IsNotifiable,
                                gatt_character.contents.character[j].HasExtendedProperties), self.scr)
                            print("  [B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
                                gatt_character.contents.character[j].IsBroadcastable,
                                gatt_character.contents.character[j].IsReadable,
                                gatt_character.contents.character[j].IsWritable,
                                gatt_character.contents.character[j].IsSignedWritable,
                                gatt_character.contents.character[j].IsWritableWithoutRespon,
                                gatt_character.contents.character[j].IsIndicatable,
                                gatt_character.contents.character[j].IsNotifiable,
                                gatt_character.contents.character[j].HasExtendedProperties))

                            gatt_character_val = self.sdk_dll.GATT_GetCharacteristicsValue(c_ubyte(j + 1))
                            if gatt_character_val.contents is not None and gatt_character.contents.character[j].IsReadable == 1:
                                try:
                                    for s in range(gatt_character_val.contents.DataSize):
                                        self.log.write("   value:0x%02x" % gatt_character_val.contents.Data[s], self.scr)
                                        print("   value:0x%02x" % gatt_character_val.contents.Data[s], end='')
                                    print('')
                                except:
                                    print('')
                            gatt_des = self.sdk_dll.GATT_GetDescriptors(c_ubyte(j + 1))
                            try:
                                if gatt_des.contents is not None:
                                    for n in range(gatt_des.contents.num):
                                        uuid = gatt_des.contents.des[i].DescriptorUuid.ShortUuid
                                        src = self.sdk_dll.ATT_GETUUIDType(uuid)
                                        self.log.write("    Descriptor id:%d,[Hdl,D:0x%04x] %s" % (
                                            n, gatt_des.contents.des[i].AttributeHandle, src.decode()),self.scr)
                                        print("    Descriptor id:%d,[Hdl,D:0x%04x] %s" % (
                                            n, gatt_des.contents.des[i].AttributeHandle, src.decode()))
                                        self.log.write("    character handle =" +
                                              gatt_des.contents.des[i].CharacteristicHandle, self.scr)
                                        print("    character handle =" +
                                              gatt_des.contents.des[i].CharacteristicHandle)

                                        gatt_des_val = self.sdk_dll.GATT_GetDescriptorValue(c_ubyte(n + 1))
                                        if gatt_des_val is not None:
                                            # print('             gatt_des_val')
                                            if gatt_des_val.contents.DescriptorType == 2:
                                                print("    Notify:%d,Ind:%d" % (
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToNotification,
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToIndication))
                            except Exception as e:
                                print('', end='')
            return 'pass'
        except Exception as e:
            print(e)
            return 'Fail'

    def Gatt_Disconnection(self, handle):
        self.log.write('断开GATT连接', self.scr)
        self.sdk_dll.GATT_EndSession.argtype = [c_ulong]
        self.sdk_dll.GATT_EndSession(handle)
        self.conn_num_gatt -= 1
        self.hConn_info_gatt[self.devices_id]['gatt'] = 0
        self.log.write('当前GATT连接数量'+str(self.conn_num_gatt), self.scr, mode=1)
        self.sdk_dll.GATT_DeInit()

    def Gatt_Connection(self, braddr, brclass=0, pin='1234', mode=13):
        addr = self.str_to_addr(braddr)
        CALLBACK = CFUNCTYPE(None, c_ulong, POINTER(c_ubyte))
        callBackFunc = CALLBACK(self.SearchCompleteCback)
        self.sdk_dll.SearchCompleCallback(callBackFunc)
        self.sdk_dll.BluetSearchRemoteDevice.argtype = [c_ulong, c_char_p]
        self.sdk_dll.BluetSearchRemoteDevice.restype = c_ulong
        brclass = c_ulong(int(brclass))

        self.sdk_dll.BluetSetLocalMode.argtypes = [c_ushort]
        mode = c_ushort(mode)
        self.sdk_dll.BluetSetLocalMode(mode)  # 设置手机蓝牙发现模式

        self.sdk_dll.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
        pin_code = c_char_p(bytes(pin, 'utf-8'))
        pin_code_len = len(pin)
        self.sdk_dll.BluetSetFixedPinCode(pin_code, pin_code_len)
        if addr == 0:
            print('连接设备的蓝牙地址错误！')
            return 0

        bdaddr = (c_char * 6)(*addr)
        Search_flag = self.sdk_dll.BluetSearchRemoteDevice(brclass, bdaddr)
        print('模块搜索中！')
        time_flag = 0
        while True:
            time.sleep(2)
            time_flag += 2
            if time_flag >= 20:
                return 0
            if self.search_flag == 1:
                self.search_flag = 0
                print('已经搜索到BLE设备')
                self.sdk_dll.BluetStopDiscoverDevice()
                print('停止搜索')
                print(self.dev_hdl)
                self.sdk_dll.GATT_Connect.argtypes = [c_ulong]
                self.sdk_dll.GATT_Initial()
                flag = self.sdk_dll.GATT_Connect(self.dev_hdl)
                if flag == 0:
                    if self.devices_id+1 > len(self.hConn_info_gatt):
                        dev_dic = {}
                        dev_dic['id'] = self.devices_id
                        dev_dic['dev_hdl'] = self.dev_hdl
                        dev_dic['bd'] = braddr
                        dev_dic['gatt'] = 1
                        self.hConn_info_gatt.append(dev_dic)
                    else:
                        self.hConn_info_gatt[self.devices_id]['dev_hdl'] = self.dev_hdl
                        self.hConn_info_gatt[self.devices_id]['bd'] = braddr
                        self.hConn_info_gatt[self.devices_id]['gatt'] = 1
                    print('模块GATT连接成功！')
                    self.conn_num_gatt += 1
                    print('当前GATT连接数量：', end='')
                    print(self.conn_num_gatt)
                    return 1
                else:
                    print('模块GATT连接失败')
                    return 0

    def simple_conn(self, con_handle, times, flags):
        '''
        测试多连接
        :param con_handle:
        :param times:
        :return:
        '''
        conn_success_flag = 0
        for i in range(1, times + 1):
            time.sleep(3)
            conn = self.sdk_dll.BluetConnectRemoteDevice(con_handle)
            if conn == 13 or conn == 0:
                self.log.write('第' + str(i + 1 + flags) + '次连接失败！', self.scr)
            else:
                self.conn_num_spp += 1
                conn_success_flag += 1
                self.log.write('第' + str(i + 1 + flags) + '次连接成功！', self.scr)
                time.sleep(2)
                self.Disconnection(conn)
        return conn_success_flag

    def gatt_simple_conn(self, dev_handle, times, flags):
        conn_success_flag = 0
        for i in range(1, times + 1):
            self.sdk_dll.GATT_Initial()
            time.sleep(2)
            conn = self.sdk_dll.GATT_Connect(dev_handle)
            if conn == 1:
                self.log.write('第' + str(i + 1 + flags) + '次连接失败！', self.scr)
            else:
                self.conn_num_gatt += 1
                conn_success_flag += 1
                self.log.write('第' + str(i + 1 + flags) + '次连接成功！', self.scr)
                time.sleep(2)
                self.Gatt_Disconnection(dev_handle)
        return conn_success_flag

    def master_read_data(self):  # 手机收
        self.sdk_dll.BluetGetRemoteBuff.restype = c_char_p
        res = self.sdk_dll.BluetGetRemoteBuff()
        return res.decode('utf-8', 'ignore').replace('\r', '').replace('\n', '')

    def salve_read_data(self, length=200, timeout=4000):  # 模块收
        ReadLength = c_int(length)
        ReadData1 = create_string_buffer(length)
        self.pDll_su.ATTestReadSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        self.pDll_su.ATTestReadSerial.restype = c_bool
        read_flag = self.pDll_su.ATTestReadSerial(ReadData1, ReadLength, timeout)
        if read_flag:
            data = ReadData1.raw.decode('utf-8', 'ignore')
            result = ''.join(re.findall(r'[A-Za-z0-9_:]|[\s]|[\n]|[-]', data))
            return result.replace('\r', '').replace('\n', ' ')
        else:
            return 0

    def master_write_data(self, hConn, data, length):  # 手机发
        self.sdk_dll.BluetWriteToRemoteDevice.argtype = [c_ulong, c_char_p, c_ushort]
        self.sdk_dll.BluetWriteToRemoteDevice(hConn, data, length)

    def salve_write_data(self, data, timeout=4000):  # 模块发
        WriteLength = c_int(len(data))
        self.pDll_su.ATTestWriteSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        self.pDll_su.ATTestWriteSerial(bytes(data, 'utf-8'), WriteLength, timeout)

    def order_analysis(self, order_list):
        '''
        指令解析
        :param order_list:
        :return: dict
        '''
        od_list = []
        for i in range(len(order_list)):
            if i is None:
                continue
            org = order_list[i].strip().split(' ')
            cmd_dic = {'type': org[0]}
            for reg in range(1, len(org), 2):
                if len(org[reg]) == 1:
                    cmd_dic[org[reg]] = org[reg + 1]
                else:
                    cmd_dic[org[reg][1:]] = org[reg + 1]
            od_list.append(cmd_dic)
        return od_list

    def flow_control(self, state):
        self.pDll_su.SetSerialRtsState.argtype = [c_bool]
        self.pDll_su.SetSerialRtsState.retype = c_bool
        flag = self.pDll_su.SetSerialRtsState(c_bool(state))
        if self.flow_control_flag == 1 and int(state) == 0:
            self.flow_control_flag = 0
        if self.flow_control_flag == 0 and int(state) == 1:
            self.flow_control_flag = 1
        if flag == 1 and int(state) == 1:
            print('流控开启成功')
        if flag == 0 and int(state) == 1:
            print('流控开启失败')
        if flag == 1 and int(state) == 0:
            print('流控关闭成功')
        if flag == 0 and int(state) == 0:
            print('流控关闭失败')

    def gatt_send_data(self, length, leave, handle, mtu=100):
        self.sdk_dll.GATT_GetFF00.argtype = [c_ulong]
        self.sdk_dll.GATT_GetFF00.restype = c_ubyte
        flag = self.sdk_dll.GATT_GetFF00(handle)
        time.sleep(2)
        if mtu > 20:
            self.sdk_dll.GATT_ExchangeMTU.argtype = [c_ulong, c_ushort]
            self.sdk_dll.GATT_ExchangeMTU.restype = c_long
            res = self.sdk_dll.GATT_ExchangeMTU(c_ulong(self.hConn_info_gatt[self.devices_id]['dev_hdl']), c_ushort(mtu))
        self.sdk_dll.GATT_ConnectFF00service.argtype = [c_ulong]
        self.sdk_dll.GATT_ConnectFF00service(handle)
        self.sdk_dll.GATT_SendMessage2FF02.argtype = [c_ulong, c_char_p]
        self.sdk_dll.GATT_SendMessage2FF02.restype = c_ubyte
        data = ''.center(mtu, 's')
        read_length = 0
        flag_time = 0
        self.salve_read_data(timeout=4000)  # 清空缓存
        while flag_time < length:
            res = self.sdk_dll.GATT_SendMessage2FF02(c_ulong(handle), c_char_p(bytes(data, 'utf-8')))
            time.sleep(1)
            if res == 0:
                read = self.salve_read_data(timeout=8000)
                if read != 0:
                    read_length += len(read)+3
                flag_time += 1
        data = ''.center(leave, 's')
        flag_time = 0
        while flag_time < 10:
            res = self.sdk_dll.GATT_SendMessage2FF02(c_ulong(handle), c_char_p(bytes(data, 'utf-8')))
            time.sleep(0.5)
            if res == 0:
                read = self.salve_read_data(timeout=8000)
                if read != 0:
                    if leave < mtu - 3:
                        read_length += len(read)
                    else:
                        read_length += len(read)+3
                break
            flag_time += 2
        if read_length == 0:
            print('未收到数据')
            return '手机未收到数据!'
        if read_length == mtu * length + leave:
            return 'pass'
        if read_length > mtu * length + leave:
            print('数据增加')
            return '数据增加' + str(read_length - (mtu * length + leave))
        if read_length < mtu * length + leave:
            print('数据丢失')
            return '数据丢失' + str(mtu * length + leave - read_length)

    def data_client2server(self, dl, handle, mtu=20):   # dangle与dangle之间
        self.sdk_dll.GATT_ConnectFF00service()
        self.sdk_dll.GATT_SendMessage2FF02.argtypes = [c_char_p]
        self.sdk_dll.GATT_SendMessage2FF02.restype = c_ubyte
        self.sdk_dll.GATT_GetCredit.restype = c_ubyte
        self.sdk_dll.NotifiDataDeal.restype = ServerStruct
        if mtu > 20:
            self.sdk_dll.GATT_ExchangeMTU.argtypes = [c_ulong, c_ushort]
            self.sdk_dll.GATT_ExchangeMTU.restype = c_long
            mtu = c_ushort(mtu)
            res = self.sdk_dll.GATT_ExchangeMTU(self.hConn_info_gatt[self.devices_id]['dev_hdl'], mtu)
        length = int(dl/mtu-2)
        leave = dl % mtu
        data = ''.center(mtu, 's')
        read_length = 0
        flag_time = 0
        while flag_time < length:
            res = self.sdk_dll.GATT_SendMessage2FF02(c_char_p(bytes(data, 'utf-8')))
            time.sleep(0.1)
            if res == 0:
                read = self.self.sdk_dll.NotifiDataDeal()
                read_length += read.data_size
                flag_time += 1
        data = ''.center(leave, 's')
        flag_time = 0
        while flag_time < 10:
            res = self.sdk_dll.GATT_SendMessage2FF02(c_char_p(bytes(data, 'utf-8')))
            time.sleep(0.1)
            if res == 0:
                read = self.self.sdk_dll.NotifiDataDeal()
                read_length += read.data_size
                break
            flag_time += 2
        if read_length == 0:
            return '手机未收到数据!'
        if read_length == dl:
            return 'pass'
        elif read_length > dl:
            return '数据增加' + str(read_length - dl)
        elif read_length < dl:
            return '数据丢失' + str(dl - read_length)

    def data_server2client(self, dl, handle, mtu=20):
        self.sdk_dll.NotifiDataSendToClient.argtype = [c_char_p, c_ushort]
        self.sdk_dll.GATT_GetFF01Data.restype = c_char_p
        self.sdk_dll.GATT_GetFF01Data.argtype = [c_ulong]
        length = int(dl / mtu)
        leave = dl % mtu
        data = ''.center(mtu, 's')
        read_length = 0
        for i in range(length):
            self.sdk_dll.NotifiDataSendToClient(data, len(data))
            time.sleep(0.1)
            read_data = self.sdk_dll.GATT_GetFF01Data(handle)
            read_data = read_data.decode('utf-8', 'ignore')
            if 'no str' not in read_data:
                read_length += len(read_data)
        data = ''.center(leave, 's')
        self.sdk_dll.NotifiDataSendToClient(data, len(data))
        time.sleep(0.1)
        read_data = self.sdk_dll.GATT_GetFF01Data(handle)
        read_data = read_data.decode('utf-8', 'ignore')
        if 'no str' not in read_data:
            read_length += len(read_data)
        if read_length == 0:
            return '手机未收到数据!'
        if read_length == dl:
            return 'pass'
        elif read_length > dl:
            return '数据增加' + str(read_length - dl)
        elif read_length < dl:
            return '数据丢失' + str(dl - read_length)

    # 多连接时，需要修改逻辑,修改连接上了就变为1（可以直接通信），连接上但是不能通信为：2
    # 每一次发送数据，都需要遍历连接信息列表，如果为2就变为1，如果为1就不变，一个列表里面只能有一个1，其余全为0或者2
    def soc_del(self, order, at_cmd=None, data=None, dt=None, dl=None, fc=None, MTU=50, mtu=20, t=None):  # AT指令
        '''
        处理soc命令
        :param mtu:
        :param t:
        :param fc:
        :param MTU:
        :param order: soc命令
        :param at_cmd: AT指令
        :param data: 发送的数据
        :param dt: 发送数据的类型：int\string
        :param dl: 发送数据的长度
        :return:
        '''
        if at_cmd is not None:
            print('发送AT指令：' + at_cmd)
            self.salve_read_data(timeout=6000)  # 清空模块缓存
            self.salve_write_data(at_cmd + '\r')
            read = self.salve_read_data(timeout=6000)
            if self.flow_control_flag == 1 or fc == 1:
                if read == 0:
                    return 'pass'
                else:
                    return '流控开启收到了数据'
            if read == 0:
                return '模块没有响应'
            print("读取：" + str(read))
            return str(read)
        else:
            if data is not None:
                print('发送数据：' + data)
                read_data = ' '
                if t == 'spp' or t == 'SPP':
                    if self.conn_num_spp <= 0:
                        print('未建立spp连接')
                        return '未建立SPP连接'
                    for i in range(self.conn_num_spp):
                        if self.hConn_info[i]['spp'] == 1:
                            self.salve_write_data(data)
                            # time.sleep(0.1)
                            read_data = self.master_read_data()
                if t == 'gatt' or t == 'GATT':
                    if self.conn_num_gatt <= 0:
                        print('未建立gatt连接')
                        return '未建立GATT连接'
                    self.sdk_dll.GATT_GetFF01Data.restype = c_char_p
                    self.sdk_dll.GATT_GetFF01Data.argtype = [c_ulong]
                    for i in range(self.conn_num_gatt):
                        if self.hConn_info_gatt[i]['gatt'] == 1:
                            self.salve_write_data(data)
                            time.sleep(0.1)
                            read_data = self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[i]['dev_hdl'])
                            read_data = read_data.decode('utf-8', 'ignore')
                if self.flow_control_flag == 1 or fc == 1:
                    if 'no str' in read_data:
                        return 'pass'
                    elif read_data == ' ':
                        return '指令错误，未指定通信方式'
                    else:
                        return '流控开启收到了数据'
                if read_data == 'no str':
                    return '手机未收到数据'
                elif read_data == ' ':
                    return '指令错误，未指定通信方式'
                else:
                    print('手机收到数据：' + read_data)
                    if read_data == data:
                        return 'pass'
                    return '数据丢失！'
            else:
                if dt is not None and dl is not None:
                    if t == 'SPP' or t == 'spp':
                        if self.conn_num_spp <= 0:
                            print('未建立spp连接')
                            return '未建立SPP连接'
                        for s in range(self.conn_num_spp):            # 多连接时，需要修改逻辑
                            if self.hConn_info[s]['spp'] == 1:
                                read_length = 0
                                self.master_read_data()
                                if dt == 'STRING' or dt == 'string':
                                    if int(fc) == 0 and self.flow_control_flag == 0:
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(MTU, 's')
                                        length = int(int(dl) / MTU)
                                        leave = int(dl) % MTU
                                        self.master_read_data()
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.master_read_data()
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.master_read_data()
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        if read_length == 0:
                                            return '手机未收到数据!'
                                        if read_length == int(dl):
                                            return 'pass'
                                        elif read_length > int(dl):
                                            return '数据增加' + str(read_length - int(dl))
                                        elif read_length < int(dl):
                                            return '数据丢失' + str(int(dl) - read_length)
                                    elif int(fc) == 1 or self.flow_control_flag == 1:
                                        print('开启流控')
                                        self.flow_control(int(fc))
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(MTU, 's')
                                        length = int(int(dl) / MTU)
                                        leave = int(dl) % MTU
                                        self.master_read_data()
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.master_read_data()
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.master_read_data()
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        print('手机收到数据：' + str(read_length))
                                        if read_length == 0:
                                            return 'pass'
                                        else:
                                            return '流控开启，手机收到了' + str(read_length) + '字节的数据'
                                if dt == 'INT' or dt == 'int':
                                    if int(fc) == 0 and self.flow_control_flag == 0:
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(MTU, '2')
                                        length = int(int(dl) / MTU)
                                        leave = int(dl) % MTU
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.master_read_data()
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.master_read_data()
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        if read_length == 0:
                                            return '手机未收到数据!'
                                        if read_length == int(dl):
                                            return 'pass'
                                        elif read_length > int(dl):
                                            return '数据增加' + str(read_length - int(dl))

                                        elif read_length < int(dl):
                                            return '数据丢失' + str(int(dl) - read_length)
                                    elif int(fc) == 1 or self.flow_control_flag == 1:
                                        print('开启流控')
                                        self.flow_control(int(fc))
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(MTU, '1')
                                        length = int(int(dl) / MTU)
                                        leave = int(dl) % MTU
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.master_read_data()
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.master_read_data()
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        print('手机收到数据：' + str(read_length))
                                        if read_length == 0:
                                            return 'pass'
                                        else:
                                            return '流控开启收到了' + str(read_length) + '字节的数据'
                    elif t == 'GATT' or t == 'gatt':
                        if self.conn_num_gatt <= 0:
                            print('未建立gatt连接')
                            return '未建立GATT连接'
                        for s in range(self.conn_num_gatt):
                            if self.hConn_info_gatt[s]['gatt'] == 1:
                                self.sdk_dll.GATT_ConnectFF00service.argtype = [c_ulong]
                                time.sleep(2)
                                self.sdk_dll.GATT_GetFF00.argtype = [c_ulong]
                                self.sdk_dll.GATT_GetFF00.restype = c_ubyte
                                flag = self.sdk_dll.GATT_GetFF00(self.hConn_info_gatt[s]['dev_hdl'])
                                self.sdk_dll.GATT_ConnectFF00service(self.hConn_info_gatt[s]['dev_hdl'])  # 打开notify
                                self.sdk_dll.GATT_GetFF01Data.restype = c_char_p
                                self.sdk_dll.GATT_GetFF01Data.argtype = [c_ulong]
                                self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[s]['dev_hdl'])
                                read_length = 0

                                if dt == 'STRING' or dt == 'string':
                                    if int(fc) == 0 and self.flow_control_flag == 0:
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(mtu, 's')
                                        length = int(int(dl) / mtu)
                                        leave = int(dl) % mtu
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[s]['dev_hdl'])
                                            read_data = read_data.decode('utf-8', 'ignore')
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[s]['dev_hdl'])
                                        read_data = read_data.decode('utf-8', 'ignore')
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        if read_length == 0:
                                            return '手机未收到数据!'
                                        if read_length == int(dl):
                                            return 'pass'
                                        elif read_length > int(dl):
                                            return '数据增加' + str(read_length - int(dl))
                                        elif read_length < int(dl):
                                            return '数据丢失' + str(int(dl) - read_length)
                                    elif int(fc) == 1 or self.flow_control_flag == 1:
                                        print('开启流控')
                                        self.flow_control(int(fc))
                                        time.sleep(0.5)
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center(mtu-2, 's')
                                        length = int(int(dl) / (mtu-2))
                                        leave = int(dl) % (mtu-2)
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[s]['dev_hdl'])
                                            read_data = read_data.decode('utf-8', 'ignore')
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.sdk_dll.GATT_GetFF01Data(self.hConn_info_gatt[s]['dev_hdl'])
                                        read_data = read_data.decode('utf-8', 'ignore')
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        print('手机收到数据：' + str(read_length))
                                        if read_length == 0:
                                            return 'pass'
                                        else:
                                            return '流控开启，手机收到了' + str(read_length) + '字节的数据'
                                if dt == 'INT' or dt == 'int':
                                    if int(fc) == 0 and self.flow_control_flag == 0:
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center((mtu-2), '2')
                                        length = int(int(dl) / (mtu-2))
                                        leave = int(dl) % (mtu-2)
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            read_data = self.sdk_dll.GATT_GetFF01Data()
                                            read_data = read_data.decode('utf-8', 'ignore')
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.sdk_dll.GATT_GetFF01Data()
                                        read_data = read_data.decode('utf-8', 'ignore')
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        if read_length == 0:
                                            return '手机未收到数据!'
                                        if read_length == int(dl):
                                            return 'pass'
                                        elif read_length > int(dl):
                                            return '数据增加' + str(read_length - int(dl))

                                        elif read_length < int(dl):
                                            return '数据丢失' + str(int(dl) - read_length)
                                    elif int(fc) == 1 or self.flow_control_flag == 1:
                                        print('开启流控')
                                        self.flow_control(int(fc))
                                        print('模块发送数据: ' + str(dl) + '位')
                                        data = ''.center((mtu-2), '1')
                                        length = int(int(dl) / (mtu-2))
                                        leave = int(dl) % (mtu-2)
                                        for i in range(length):
                                            self.salve_write_data(data)
                                            time.sleep(0.1)
                                            read_data = self.sdk_dll.GATT_GetFF01Data()
                                            read_data = read_data.decode('utf-8', 'ignore')
                                            if 'no str' not in read_data:
                                                read_length += len(read_data)
                                        self.salve_write_data(''.center(leave, 's'))
                                        time.sleep(0.1)
                                        read_data = self.sdk_dll.GATT_GetFF01Data()
                                        read_data = read_data.decode('utf-8', 'ignore')
                                        if 'no str' not in read_data:
                                            read_length += len(read_data)
                                        print('手机收到数据：' + str(read_length))
                                        if read_length == 0:
                                            return 'pass'
                                        else:
                                            return '流控开启收到了' + str(read_length) + '字节的数据'
                    else:
                        return '未指定通信方式'
                else:
                    if fc is not None:
                        self.flow_control(int(fc))
                        time.sleep(2)
                        return 'pass'
                    else:
                        return 'soc指令错误'
        return 'soc指令错误'

    def remote_del(self, order, device_id='0', cmd=None, bd=None, bt='public', t=None, dl=None, cf=None, MTU=100, mtu=50, p='1234', ci=None, r=None):
        '''
        处理remote命令
        :param ci:
        :param mtu:
        :param MTU:
        :param cf:
        :param dl:
        :param order: remote命令
        :param device_id: remote id
        :param cmd: 执行命令
        :param bd: 蓝牙地址
        :param bt: 蓝牙地址类型
        :param t: 蓝牙连接类型
        :param p:
        :return: pass\Fail
        '''
        if cmd == 'inquiry' or cmd == 'INQUIRY':
            if t == 'GATT' or t == 'GATT':
                if int(device_id) < len(self.hConn_info_gatt):
                    if self.hConn_info_gatt[int(device_id)]['gatt'] > 0:
                        flag = self.Get_Gatt_Services(self.hConn_info_gatt[int(device_id)]['dev_hdl'])
                        return flag
                return '未建立gatt连接'
            print('本机信息：')
            self.sdk_dll.BluetGetLocalInfo.restype = LocalInfo
            time.sleep(2)
            self.local_info = self.sdk_dll.BluetGetLocalInfo()
            print('设备名称：' + str(self.local_info.name.decode('utf-8')))
            print('设备类型：' + str(self.local_info.device_class))
            addr = ''
            for i in range(len(self.local_info.bd_addr)):
                if len(hex(self.local_info.bd_addr[i])) == 3:
                    addr += '0' + hex(self.local_info.bd_addr[i])[-1]
                if len(hex(self.local_info.bd_addr[i])) == 4:
                    addr += hex(self.local_info.bd_addr[i])[-2:]
            print('模块地址：' + addr)
            print('设备发现模式：' + str(self.local_info.discover_mode))
            return 'pass'

        elif cmd == 'CONN' or cmd == 'conn':
            if bd is not None:
                if bt == 'public' or bt == 'PUBLIC':
                    print('地址类型：public ')
                elif bt == 'random' or bt == 'RANDOM':
                    print('地址类型：random')
                if t == 'SPP' or t == 'spp':
                    print('进行spp连接')
                    self.log.write('进行spp连接', self.scr)
                    self.init_spp_service(1)
                    flag = self.Connection(braddr=bd,  pin=p)
                    time.sleep(2)
                    if flag:
                        return 'pass'
                    return 'spp连接失败'
                elif t == 'GATT' or t == 'gatt':
                    if ci is not None:
                        if ci.isdigit():
                            if int(ci) < len(self.hConn_info_gatt):
                                if int(ci) == int(device_id):
                                    return 'remote连接指令错误'
                                flag = self.Gatt_Connection(self.device_info[int(ci)]['bd'], pin=p)   # 连接dangle
                                if flag:
                                    return 'pass'
                                return 'gatt连接失败'
                    print('进行GATT连接')
                    self.log.write('进行GATT连接', self.scr)
                    flag = self.Gatt_Connection(braddr=bd,  pin=p)
                    time.sleep(2)
                    if flag:
                        return 'pass'
                    return 'gatt连接失败'
                elif t is None:
                    return '没有选择连接方式'
                else:
                    return '指令错误！'
            else:
                print('蓝牙地址为空')
                return '蓝牙地址为空'

        elif cmd == 'disc' or cmd == 'DISC':
            if t == 'SPP' or t == 'spp':
                if len(self.hConn_info) > int(device_id):
                    if self.hConn_info[int(device_id)]['spp'] == 1:
                        time.sleep(1)
                        self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'])
                        time.sleep(2)
                        return 'pass'
                return '未建立SPP连接'
            elif t == 'GATT' or t == 'gatt':
                print('断开GATT连接')
                # self.log.write('断开GATT连接', self.scr)
                if len(self.hConn_info_gatt) > int(device_id):
                    if self.hConn_info_gatt[int(device_id)]['gatt'] == 1:
                        # print(self.hConn_info_gatt[int(device_id)]['dev_hdl'])
                        time.sleep(2)
                        self.Gatt_Disconnection(self.hConn_info_gatt[int(device_id)]['dev_hdl'])
                        time.sleep(1)
                        return 'pass'
                return '未建立GATT连接'
            else:
                return '未指定断开连接的方式！'

        elif cmd == 'send' or cmd == 'SEND':
            if int(dl) > 0:
                data = ''.center(MTU, 's')
                self.salve_read_data()  # 清空模块的缓存区
                data_length = 0
                if t == 'GATT' or t == 'gatt':
                    length = int(int(dl) / mtu)
                    leave = int(dl) % mtu
                    if r == 'c' or r == 'C':
                        return self.data_client2server(int(dl), mtu=mtu)
                    elif r == 'S' or r == 's':
                        return self.data_server2client(int(dl), mtu=mtu)
                    elif len(self.hConn_info_gatt) > int(device_id):
                        if self.hConn_info_gatt[int(device_id)]['gatt'] == 1:
                            handle = self.hConn_info_gatt[int(device_id)]['dev_hdl']
                            result = self.gatt_send_data(length, leave, handle, mtu)
                            return result
                    else:
                        return '指令错误'
                elif t == 'SPP' or t == 'spp':
                    length = int(int(dl) / MTU)
                    leave = int(dl) % MTU
                    if r == 'c' or r == 'C':
                        pass
                    elif r == 's' or r == 'S':
                        pass
                    elif len(self.hConn_info) > int(device_id):
                        if self.hConn_info[int(device_id)]['spp'] == 1:
                            self.salve_read_data()
                            for i in range(length):
                                str_len = c_ushort(len(data))
                                str_p = c_char_p(bytes(data, 'utf-8'))
                                self.master_write_data(self.hConn_info[int(device_id)]['conn_hdl'], str_p, str_len)
                                read_data = self.salve_read_data(timeout=8000)
                                if read_data != 0:
                                    data_length += len(read_data)
                        else:
                            return '未建立连接'
                    else:
                        return '未建立连接'
                    data = ''.center(leave, 's')
                    str_len = c_ushort(len(data))
                    str_p = c_char_p(bytes(data, 'utf-8'))
                    self.master_write_data(self.hConn_info[int(device_id)]['conn_hdl'], str_p, str_len)
                    read_data = self.salve_read_data(timeout=8000)
                    if read_data != 0:
                        data_length += len(read_data)
                    if data_length == 0:
                        return '模块没有收到数据'
                    if data_length == int(dl):
                        return 'pass'
                    elif data_length > int(dl):
                        return '数据增加' + str(data_length - int(dl))
                    elif data_length < int(dl):
                        return '数据丢失' + str(int(dl) - data_length)
                else:
                    return '指令错误！'
            else:
                return 'SEND指令错误'

        elif cmd == 'recv' or cmd == 'RECV':
            master_read = self.master_read_data()
            if len(master_read) > 0:
                return 'pass'
            return '未收到数据'

        elif cmd == 'DISCS' or cmd == 'discs':
            if cf is not None:
                if int(cf) > 0:
                    times = int(cf)
                else:
                    return 'DISCS指令错误'
            else:
                return 'DISCS指令错误'
            success_times = 0
            if t == 'SPP' or t == 'spp':
                self.init_spp_service(1)
                flag = self.Connection(braddr=bd)
                if flag:
                    self.log.write('第1次连接成功', self.scr)
                    time.sleep(2)
                    self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'], del_hdl=0)
                    success_times = 1
                    success_times += self.simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - flag, 0)
                else:
                    self.log.write('第1次连接失败', self.scr)
                    for i in range(1, times):
                        flag = self.Connection(braddr=bd)
                        if flag:
                            self.log.write('第' + str(i + 1) + '次连接成功', self.scr)
                            self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'])
                            success_times = 1
                            success_times += self.simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - i - 1, i-1)
                            break
                        else:
                            self.log.write('第' + str(i + 1) + '次连接失败', self.scr)
                if success_times == times:
                    return 'pass'
                return '连接'+str(times)+'次' + '成功' + str(success_times)+'次'
            if t == 'GATT' or t == 'gatt':
                flag = self.Gatt_Connection(braddr=bd)
                if flag:
                    self.log.write('第1次连接成功', self.scr)
                    self.Gatt_Disconnection(self.hConn_info_gatt[int(device_id)]['dev_hdl'])
                    success_times = flag
                    success_times += self.gatt_simple_conn(self.hConn_info_gatt[int(device_id)]['dev_hdl'], times - flag, 0)
                else:
                    self.log.write('第1次连接失败', self.scr)
                    for i in range(1, times):
                        flag = self.Gatt_Connection(braddr=bd)
                        if flag:
                            self.log.write('第'+str(i+1)+'次连接成功', self.scr)
                            self.Gatt_Disconnection(self.hConn_info[int(device_id)]['dev_hdl'])
                            success_times = 1
                            success_times += self.gatt_simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times-i-1, i-1)
                            break
                        else:
                            self.log.write('第' + str(i + 1) + '次连接失败', self.scr)
                if success_times == times:
                    return 'pass'
                return '连接' + str(times) + '次' + '成功' + str(success_times) + '次'

        else:
            return 'remote指令错误'

    def local_del(self, order, cmd=None, bd=None, bt='public', t=None):  # 待拓展
        '''

        :param order:
        :param cmd:
        :param bd:
        :param bt:
        :param t:
        :return:
        '''
        return 'pass'

    def order_del_list(self, order_list):
        result_list_tmp = []
        # time.sleep(0.1)
        c = None
        d = None
        dt = None
        dl = None
        bd = None
        fc = '0'
        bt = 'public'
        t = None
        id = '0'
        result = ''
        cf = None
        p = '1234'
        ci = None
        r = None
        order_del_list = self.order_analysis(order_list)
        for order in order_del_list:
            print(order)
            # try:
            if order['type'] == 'soc' or order['type'] == 'SOC':
                if 'C' in order.keys():
                    c = order['C']
                if 'D' in order.keys():
                    d = order['D']
                if 'DT' in order.keys():
                    dt = order['DT']
                if 'DL' in order.keys():
                    dl = order['DL']
                if 'T' in order.keys():
                    t = order['T']
                if 't' in order.keys():
                    t = order['t']
                if 'c' in order.keys():
                    c = order['c']
                if 'd' in order.keys():
                    d = order['d']
                if 'dt' in order.keys():
                    dt = order['dt']
                if 'dl' in order.keys():
                    dl = order['dl']
                if 'fc' in order.keys():
                    fc = order['fc']
                if 'FC' in order.keys():
                    fc = order['FC']
                result = self.soc_del(order, at_cmd=c, data=d, dt=dt, dl=dl, fc=fc, t=t)
                c = None
                d = None
                dt = None
                dl = None
                fc = '0'
                t = None
                if result is not None and result != ' ' and result != 'pass':
                    self.log.write(result, self.scr)
                else:
                    pass
            elif order['type'] == 'remote' or order['type'] == 'REMOTE':
                if 'I' in order.keys():
                    id = order['I']
                if 'C' in order.keys():
                    c = order['C']
                if 'BD' in order.keys():
                    bd = order['BD']
                if 'BT' in order.keys():
                    bt = order['BT']
                if 'T' in order.keys():
                    t = order['T']
                if 'DL' in order.keys():
                    dl = order['DL']
                if 'CF' in order.keys():
                    cf = order['CF']
                if 'P' in order.keys():
                    p = order['P']
                if 'R' in order.keys():
                    r = order['R']
                if 'CI' in order.keys():
                    ci = order['CI']
                if 'ci' in order.keys():
                    ci = order['ci']
                if 'r' in order.keys():
                    r = order['r']
                if 'p' in order.keys():
                    p = order['p']
                if 'i' in order.keys():
                    id = order['i']
                if 'c' in order.keys():
                    c = order['c']
                if 'bd' in order.keys():
                    bd = order['bd']
                if 'bt' in order.keys():
                    bt = order['bt']
                if 't' in order.keys():
                    t = order['t']
                if 'dl' in order.keys():
                    dl = order['dl']
                if 'cf' in order.keys():
                    cf = order['cf']
                self.devices_id = int(id)
                result = self.remote_del(order, device_id=id, cmd=c, bd=bd, bt=bt, t=t, dl=dl, cf=cf, p=p, ci=ci, r=r)
                c = None
                cf = None
                dl = None
                ci = None
                r = None
                bd = None
                bt = 'public'
                t = None
                id = '0'
                p = '1234'
            elif order['type'] == 'local' or order['type'] == 'LOCAL':
                if 'C' in order.keys() or 'c' in order.keys():  # inquiry/conn/disc/send/recv
                    c = order['C']
                if 'BD' in order.keys() or 'bd' in order.keys():
                    bd = order['BD']
                if 'BT' in order.keys() or 'bt' in order.keys():
                    bt = order['BT']
                if 'T' in order.keys() or 't' in order.keys():
                    t = order['T']
                result = self.local_del(order, cmd=c, bd=bd, bt=bt, t=t)
                c = None
                bd = None
                bt = 'public'
                t = None
            else:
                print('指令错误！')
                result_list_tmp.append('指令错误！')
            print('用例执行结果：', end='')
            print(result)
            result_list_tmp.append(result)
        time.sleep(2)
        return result_list_tmp

    def read_com(self):
        """
        创建读取串口按钮控件
        :return:
        """
        self.read = Button((self.root), text='READ_COM', relief=GROOVE, command=self.thread_choose_com, bg='green',
                           fg='white',
                           height=1,
                           width=12)

    def test_contents(self):
        """
        创建单选控件
        :return:
        """
        xVariable = StringVar()
        self.contents = ttk.Combobox(self.root, textvariable=xVariable, state='readonly', width=15)
        self.contents_label = Label(self.root, text="Port:")

    def process_bar(self, percent, start_str='', total_length=0):
        """
        进度条
        :return:   {:0>4.1f}%
        """
        bar = ''.join(["\033[1;31;41m%s\033[0m" % '   '] * int(percent * total_length)) + ''
        bar = '\r' + start_str + bar.ljust(total_length) + ' {:.2f}%'.format(percent * 100)
        print(bar, end='', flush=True)
        self.scr.insert(END, time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '→' ' {:.2f}%'.format(percent * 100) + '\n')
        self.scr.see(END)
        self.scr.update()
        if percent != 1:
            self.scr.delete("%s-1l" % INSERT, INSERT)
        else:
            self.log.write('执行完成!', self.scr)

    def thread_choose_com(self):
        t = threading.Thread(target=self.option_com)
        t.setDaemon(True)
        t.start()

    def option_com(self):
        '''
        选择串口
        :return:
        '''
        self.read.config(state='disable')
        self.get_port()
        com_list = tuple(self.com_list)
        self.contents['value'] = com_list
        if len(self.contents['value']) == 0:
            tkinter.messagebox.showinfo(title='Notice', message='没有选中串口\n')
            self.log.write('没有选中串口', self.scr)
        else:
            self.contents.current(0)
        self.read.config(state='normal')

    def selectPath(self):
        '''
        选择路径
        :return:
        '''
        path = askopenfilename(filetypes=[('Python', '*.xlsx *.xls'), ('All files', '*')])
        if len(path) > 0:
            self.test_path_entry.delete(0, END)
            self.test_path_entry.insert(END, path)
            self.log.write('测试用例路径：' + path, self.scr)
            self.sheet_name_entry['value'] = tuple(['None'])
            self.sheet_name_entry.current(0)
        self.path = path

    def IsConnCallBack(self, datatype, data):
        pass

    def read_sheet(self):
        try:
            if len(self.path) > 0:
                self.read_data = openpyxl.load_workbook(self.path)
                sheet_name = []
                for name in self.read_data:
                    sheet_name.append(name.title)
                    self.sheet_name_entry['value'] = tuple(sheet_name)
                if len(self.sheet_name_entry['value']) == 0:
                    tkinter.messagebox.showinfo(title='Notice', message='没有sheet表\n')
                else:
                    self.sheet_name_entry.current(0)
        except:
            tkinter.messagebox.showinfo(title='Notice', message='路径为空！')

    def get_test_path(self):
        """
        得到测试用例的路径
        :return:
        """
        var1 = StringVar()
        var2 = StringVar()
        self.test_path_entry = Entry(self.root, textvariable=var1, bd=2, width=40, state='normal')
        self.test_path_label = Label(self.root, text="文件路径:", height=1)
        self.test_path_button = Button(self.root, text="选择文件", command=self.selectPath, height=1)
        self.sheet_name_entry = ttk.Combobox(self.root, textvariable=var2, state='readonly', width=20)
        self.sheet_name_label = Label(self.root, text="Sheet:")
        self.sheet_name_button = Button(self.root, text="读取sheet", command=self.read_sheet , height=1)

    def simple_test_map(self):
        self.simple_test_text = Text(self.root, width=109, height=5)

    def logo(self):
        self.console_label = Label(self.root, text="Console", height=1)
        self.simple_test_logo = Label(self.root, text='测试输入框', height=1)

    def get_baud(self):
        '''
        创建BAUD控件
        :return:
        '''
        self.var2 = StringVar(value='115200')
        self.test_baud_entry = Entry((self.root), textvariable=(self.var2), bd=2, width=12, state='normal')
        self.test_baud_label = Label(self.root, text="BAUD:", height=1)

    def execute_script(self):
        '''
        运行脚本
        :return:
        '''
        self.scr.delete(1.0, END)
        self.execute_button.config(state='disable')
        if len(self.test_baud_entry.get()) == 0:
            tkinter.messagebox.showinfo(title='Notice', message='BAUD不能为空')
            self.log.write('BAUD为空!!', self.scr)
            self.execute_button.config(state='normal')
        else:
            if not self.test_baud_entry.get().isdigit():
                tkinter.messagebox.showinfo(title='Notice', message='BAUD有非法字符!!\n请重新输入!!')
                self.log.write('BAUD有非法字符!!', self.scr)
                self.execute_button.config(state='normal')
            else:
                self.BAUD = int(self.test_baud_entry.get())
                self.log.write('BAUD：' + str(self.BAUD), self.scr)
                if len(self.contents.get()) == 0:
                    tkinter.messagebox.showinfo(title='Notice', message='串口为空!!')
                    self.log.write('串口为空!!', self.scr)
                    self.execute_button.config(state='normal')
                else:
                    self.com = self.contents.get()
                    if len(self.test_path_entry.get()) > 0:
                        self.path = self.test_path_entry.get()
                        if len(self.sheet_name_entry.get()) > 0 and self.sheet_name_entry.get() != 'None':
                            self.sheet_name = self.sheet_name_entry.get()
                            t = threading.Thread(target=self.test_spp)
                            t.setDaemon(True)
                            t.start()
                        elif len(self.sheet_name_entry.get()) <= 0 or self.sheet_name_entry.get() == 'None':
                            tkinter.messagebox.showinfo(title='Notice', message='没有选择sheet表！')
                            self.log.write('没有选择sheet表！', self.scr)
                            self.execute_button.config(state='normal')

                    elif len(self.simple_test_text.get('1.0', END)) > 0:
                        test = self.simple_test_text.get('1.0', END)
                        test = test.replace('\n', '')
                        t = threading.Thread(target=self.simple_test, args=(test,))
                        t.setDaemon(True)
                        t.start()

                    elif len(self.test_path_entry.get()) <= 0:
                        tkinter.messagebox.showinfo(title='Notice', message='测试用例路径为空！')
                        self.log.write('测试用例路径为空！', self.scr)
                        self.execute_button.config(state='normal')

    def simple_test(self, text):
        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        self.pDll_su.ATTestOpenSerial.retypes = c_bool
        self.pDll_su.ATCmdTestCreateObject()
        CALLBACK = CFUNCTYPE(None, c_uint32, c_void_p)
        Stop_Search_Device = CALLBACK(self.IsConnCallBack)
        self.pDll_su.ATTestSetCallback.argtypes = [CALLBACK]
        self.pDll_su.ATTestSetCallback(Stop_Search_Device)
        open_flag = self.pDll_su.ATTestOpenSerial(int(self.com[-1]), int(self.BAUD))
        if open_flag == 0:
            self.log.write('串口打开失败', self.scr)
            self.pDll_su.ATTestCloseSerial()
            self.pDll_su.ATCmdTestDeleteObject()
            self.execute_button.config(state='normal')
            return 0
        test = text.strip().replace('\n', '')
        result = None
        if '\CR' in test:
            order = test.split('\CR')
            print(order)
            order_list = []
            for i in order:
                if 'soc' in i or 'SOC' in i or 'remote' in i or 'REMOTE' in i or 'local' in i or 'LOCAL' in i:
                    order_list.append(i)
            result = self.order_del_list(order_list)
        else:
            if 'soc' in test or 'SOC' in test or 'remote' in test or 'REMOTE' in test or 'local' in test or 'LOCAL' in test:
                order_list = [test]
                result = self.order_del_list(order_list)
        flag = 1
        for i in result:
            if i is not None and i != 'Pass' and i != 'pass':
                flag = 0
                self.log.write(i, self.scr)
        if flag:
            self.log.write('测试结果： Pass', self.scr)
        else:
            self.log.write('测试结果： Fail', self.scr)


        self.pDll_su.ATTestCloseSerial()
        self.pDll_su.ATCmdTestDeleteObject()
        self.log.write('单条case执行完毕！', self.scr)
        self.execute_button.config(state='normal')

    def execute(self):
        '''
        运行控件
        :return:
        '''
        self.execute_button = Button(self.root, text='EXECUTE', relief=GROOVE, command=self.execute_script,
                                     bg='green', fg='white',
                                     height=1,
                                     width=12)

    def console(self):
        '''
        信息输出框
        :return:
        '''
        self.scr = tkinter.scrolledtext.ScrolledText((self.root), width=95, height=20, font=('楷书', 11))

    def creat_map(self):
        '''
        创建TK界面
        :return:
        '''
        self.read_com()
        self.test_contents()
        self.get_test_path()
        self.get_baud()
        self.execute()
        self.console()
        self.logo()
        self.simple_test_map()
        self.console_label.place(x=380, y=6)
        self.simple_test_logo.place(x=380, y=350)
        self.simple_test_text.place(x=10, y=380)
        self.scr.place(x=10, y=30)
        self.test_path_entry.place(x=70, y=485)
        self.test_path_label.place(x=10, y=482)
        self.test_path_button.place(x=360, y=480)
        self.sheet_name_label.place(x=450, y=482)
        self.sheet_name_entry.place(x=500, y=485)
        self.sheet_name_button.place(x=670, y=480)
        self.read.place(x=100, y=575)
        self.execute_button.place(x=600, y=575)
        self.test_baud_entry.place(x=605, y=530)
        self.test_baud_label.place(x=555, y=530)
        self.contents_label.place(x=100, y=530)
        self.contents.place(x=140, y=530)
        self.root.mainloop()

    def test_spp(self):
        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]

        self.pDll_su.ATTestOpenSerial.retypes = c_bool
        self.pDll_su.ATCmdTestCreateObject()

        CALLBACK = CFUNCTYPE(None, c_uint32, c_void_p)
        Stop_Search_Device = CALLBACK(self.IsConnCallBack)
        self.pDll_su.ATTestSetCallback(Stop_Search_Device)

        open_flag = self.pDll_su.ATTestOpenSerial(int(self.com[-1]), int(self.BAUD))
        if open_flag == 0:
            self.log.write('串口打开失败', self.scr)
            self.pDll_su.ATTestCloseSerial()
            self.pDll_su.ATCmdTestDeleteObject()
            self.execute_button.config(state='normal')
            return 0
        result_order, result_order1 = self.del_data()
        result = []
        question = []
        success_times = 0
        self.length = len(self.data[self.ini])
        for i in range(0, len(self.data[self.type[0]])):
            self.log.write('\n\n执行case: '+self.data[self.type[0]][i], self.scr)
            ini = self.order_del_list(result_order[i])  # 前置条件############################################
            process = self.order_del_list(result_order1[i])  # 测试步骤############################################
            regs = ''
            flag = 0
            success_flag = 1
            print('条件：', end='')
            print(ini)
            for reg in ini:
                if 'pass' == reg:
                    continue
                elif reg is not None:
                    if ' ' in reg:
                        temp = reg.split(' ')
                        print(self.data[self.expecte][i])
                        for temp_reg in temp:
                            if temp_reg in self.data[self.expecte][i]:
                                pass
                            elif temp_reg is None:
                                pass
                            else:
                                flag += 1
                                regs += str(flag) + '、' + temp_reg + '\n\t'
                                success_flag = 0
                    elif reg in self.data[self.expecte][i]:
                        pass
                    else:
                        success_flag = 0
                        flag += 1
                        regs += str(flag) + '、' + reg + '\n\t'
            print('过程：', end='')
            print(process)
            for reg in process:
                if 'pass' == reg:
                    continue
                elif reg is not None:
                    if ' ' in reg:
                        temp = reg.split(' ')
                        for temp_reg in temp:
                            if temp_reg in self.data[self.expecte][i]:
                                print(temp_reg)
                                pass
                            elif temp_reg is None:
                                pass
                            else:
                                flag += 1
                                regs += str(flag) + '、' + temp_reg + '\n\t'
                                success_flag = 0
                    elif reg in self.data[self.expecte][i]:
                        pass
                    else:
                        success_flag = 0
                        flag += 1
                        regs += str(flag) + '、' + reg + '\n\t'
            if success_flag == 1:
                result.append('Pass')
                question.append(' ')
                success_times += 1
            if success_flag == 0:
                result.append('Fail')
                question.append(regs)
        self.log.write('case 通过率：%.2f%%' % ((int(success_times)/len(result))*100), self.scr)
        self.execute_button.config(state='normal')
        self.log.write('正在生成报告！', self.scr)
        self.pDll_su.ATTestCloseSerial()
        self.pDll_su.ATCmdTestDeleteObject()
        self.data = None
        self.result_data = []
        gc.collect()
        self.creat_excel(result, question)

    def creat_excel(self, result=None, question=None):
        table = self.read_data.get_sheet_by_name(self.sheet_name)
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        if result is not None:
            for i in range(1, ncolumns+1):
                if table.cell(1, i).value == self.result:
                    for j in range(2, nrows+1):
                        if result[j-2] != ' ':
                            table.cell(j, i).value = str(result[j-2])
                if table.cell(1, i).value == self.question:
                    for j in range(2, nrows+1):
                        if question[j-2] != ' ':
                            table.cell(j, i).value = str(question[j-2])
        for i in range(10):
            try:
                self.read_data.save('Test_Report.xlsx')
                if result is not None:
                    self.log.write('test_report已经生成', self.scr)
                    break
            except:
                if result is not None:
                    self.log.write('报告生成文件未关闭，请先关闭！', self.scr)
                    tkinter.messagebox.showinfo(title='Notice', message='报告生成文件未关闭！')

    def delete_ini_file(self):
        inifile1 = 'REMOTEDEVICE.INI'
        inifile2 = 'LOCALSERVICE.INI'
        if os.path.isfile(inifile1):
            os.remove(inifile1)
        if os.path.isfile(inifile2):
            os.remove(inifile2)


if __name__ == "__main__":
    tool = Auto_Test()
    tool.creat_map()