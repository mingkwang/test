from ctypes import *
import _ctypes
import os
import sys
import threading
import string
import time
import tkinter as tk
import tkinter.messagebox
import tkinter.scrolledtext
from ctypes import *
from multiprocessing import Pool
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import numpy as np
import pandas as pd
import serial
import serial.tools.list_ports
import xlrd
import xlwt
import re
import openpyxl
class Log(object):
    def __init__(self):
        self.file = './log'
        if not os.path.exists(self.file):
            os.makedirs('./log')
        self.filename = self.file + '/log-' + time.strftime('%Y-%m-%d_%H-%M-%S') + '.txt'
        self.log = open(self.filename, 'a+', encoding='UTF-8')

    def write(self, message, scr, mode=1):
        self.log = open(self.filename, 'a', encoding='UTF-8')
        self.log.write(time.strftime('%Y-%m-%d_%H-%M-%S') + ': ' + message + '\n')
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ': ' + message)
        if mode == 0:
            return 0
        scr.insert(END, time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '→' + message + '\n')
        scr.see(END)
        scr.update()
        self.close()

    def close(self):
        self.log.close()
class callback:
    def __init__(self, dll):
        self.dll = dll
        return

    def Function(self, callback, p):
        return self.dll.ccrfidDevSdkStartWork(callback, p)

    def write_AT(self, callback):
        pass
class SimpStruct(Structure):
    _fields_ = [("cls", c_ulong),  # 远程设备class
                ("dev_hdl", c_ulong),  # 远程设备的device handle
                ("name", c_char * 32),  # 远程设备名字
                ("BdAddr", c_ubyte * 6),  # 远程设备地址
                ("hConn", c_ulong),  # 远程设备connection handle
                ("placement", c_char * 10)]
class ATTestList(Structure):
    pass
ATTestList._fields_ = [("ATCmdSingleCmdStr", c_char_p),  # 定义一维数组,字符串数组
                       ("ATCmdSinglelCmdLength", c_uint16),
                       ("ATCmdSingleCmdResultStr", c_char_p),
                       ("ATCmdSingleCmdFeedbackStr", c_char_p),
                       ("ATCmdCheckoutResult", c_bool),
                       ("Next", POINTER(ATTestList))]
class LmpInfo(Structure):
    _fields_ = [("lmp_feature", c_ubyte * 8),
                ("manuf_name", c_ushort),
                ("lmp_subversion", c_ushort),
                ("lmp_version", c_ubyte),
                ("hci_version", c_ubyte),
                ("hci_revision", c_ushort),
                ("country_code", c_ubyte)]
class LocalInfo(Structure):
    _fields_ = [("name", c_char * 32),
                ("discover_mode", c_ushort),
                ("pin_code", c_char * 32),
                ("device_class", c_ulong),
                ("bd_addr", c_ubyte * 6),
                ("lmp_info", LmpInfo),
                ("app_param", c_ulong)]
class BtSdkUUIDStru(Structure):
    _fields_ = [("Data1", c_ulong),
                ("Data2", c_ushort),
                ("Data3", c_ushort),
                ("Data4", c_char * 8)]
class BtsdkGATTUUIDStru(Structure):
    _fields_ = [("IsShortUuid", c_long),
                ("ShortUuid", c_ushort),
                ("LongUuid", BtSdkUUIDStru)]
class BtsdkGATTServiceStru(Structure):
    _fields_ = [("ServiceUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort)]
class BtSdkGATTService(Structure):
    _fields_ = [("num", c_ushort),
                ("service", BtsdkGATTServiceStru * 10)]
class BtsdkGATTCharacteristicStru(Structure):
    _fields_ = [("ServiceHandle", c_ushort),
                ("CharacteristicUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort),
                ("CharacteristicValueHandle", c_ushort),
                ("IsBroadcastable", c_ulong),
                ("IsReadable", c_ulong),
                ("IsWritable", c_ulong),
                ("IsWritableWithoutRespon", c_ulong),
                ("IsSignedWritable", c_ulong),
                ("IsNotifiable", c_ulong),
                ("IsIndicatable", c_ulong),
                ("HasExtendedProperties", c_ulong)]
class BtSdkGATTCharacteristic(Structure):
    _fields_ = [("num", c_ushort),
                ("character", BtsdkGATTCharacteristicStru * 10)]
class BtsdkGATTDescriptorStru(Structure):
    _fields_ = [("ServiceHandle", c_ushort),
                ("CharacteristicHandle", c_ushort),
                ("DescriptorType", c_ulong),
                ("DescriptorUuid", BtsdkGATTUUIDStru),
                ("AttributeHandle", c_ushort)]
class BtSdkGATTDescriptor(Structure):
    _fields_ = [("num", c_ushort),
                ("des", BtsdkGATTDescriptorStru * 10)]
class CEP(Structure):
    _fields_ = [("IsReliableWriteEnabled", c_ulong),
                ("IsAuxiliariesWritable", c_ulong)]
class CCC(Structure):
    _fields_ = [("IsSubscribeToNotification", c_ulong),
                ("IsSubscribeToIndication", c_ulong)]
class SCC(Structure):
    _fields_ = [("IsBoardcast", c_ulong)]
class CF(Structure):
    _fields_ = [("Format", c_char),
                ("Exponent", c_char),
                ("Uint", BtsdkGATTUUIDStru),
                ("NameSpace", c_char),
                ("Descriptor", BtsdkGATTUUIDStru)]
class U(Union):
    _fields_ = [("CharacterExtendedProperties", CEP),
                ("ClientCharacteristicConfiguration", CCC),
                ("ServerCharacteristicConfiguration", SCC),
                ("CharacteristicFormat", CF)]
class BtsdkGATTDscriptorValueStru(Structure):
    _fields_ = [("DescriptorType", c_ulong),
                ("DescriptorUuid", BtsdkGATTUUIDStru),
                ("u", U),
                ("DataSize", c_ulong),
                ("Data", c_char * 100)]
class BtsdsGATTCharacteristicValueStru(Structure):
    _fields_ = [("DataSize", c_ulong),
                ("Data", c_ubyte * 100)]


class Auto_Test(object):

    def __init__(self):
        self.root = Tk()
        self.root.geometry('800x690+400+50')
        self.root.resizable(0, 0)
        self.root.title('Automatic TestReport generation tool')
        self.ini = '初始化条件'
        self.process = '测试过程'
        self.result = '测试结果'
        self.expecte = '期望结果'
        self.question = '问题描述'
        self.sheet_name = 'SPP'
        self.com_list = []
        self.conn_flag = 0
        self.report_path = 'test_Report.xlsx'
        self.result_data = []  # 把data转换为二维数组(没有必要)
        self.type1 = ['用例ID', '类型', '级别', '目的', '初始化条件', '测试过程', '期望结果', '测试结果', '问题描述', '概率',
                     '测试手机']  # 可以先用 xlrd 读第一行的值
        self.log = Log()
        self.search_flag = 0
        self.dev_hdl = 0
        self.hConn = 0
        self.conns = []
        self.at_res = ''
        self.isconnflag = 0
        self.isPDU = 0
        self.devices_id = 0  # remote_index
        self.conn_num_spp = 0
        self.conn_num_gatt = 0
        self.hConn_info = []  # 存储remote连接的搜索句柄，连接句柄[{id:0,dev_hdl:搜索句柄,conn_hdl:连接句柄.....},
        # {id:1,dev_hdl:搜索句柄,conn_hdl:连接句柄.....}......]
        self.sdk_dll = CDLL("sdksample.dll")
        self.init()
        self.pDll_su = CDLL("ATTestLib.dll")

    def init(self):
        self.sdk_dll.BluetDisConnectRemoteDevice.argtype = [c_ulong]
        self.sdk_dll.BluetDisConnectRemoteDevice.retypes = c_ushort
        self.sdk_dll.BluetFuncInterfaceInit()
        self.sdk_dll.Get_Version.retypes = c_char_p
        version = self.sdk_dll.Get_Version()
        print('当前版本：' + str(version))

    def SearchCompleteCback(self, dev_hdl, bd_addr):
        self.search_flag = 1
        self.dev_hdl = dev_hdl

    # def _callback(self, stype):
    #     print(stype)
    #     if stype == 1:
    #         self.search_flag = 1
    #     if stype == 2:
    #         self.sdk_dll_liu.SetRemoteDevicePinCode.argtypes = [c_char_p]
    #         code = c_char_p(bytes("1234", 'utf-8'))
    #         self.sdk_dll_liu.SetRemoteDevicePinCode(code)
    #     if stype == 3:
    #         self.sdk_dll_liu.Accept_Conncet.argtypes = [c_char_p]
    #         cmd = c_char_p(bytes("y", 'utf-8'))
    #         self.sdk_dll_liu.Accept_Conncet(cmd)
    #     if stype == 4:
    #         self.sdk_dll_liu.Get_Remote_Buff.restype = c_char_p
    #         res = self.sdk_dll_liu.Get_Remote_Buff()
    #         print('回调手机接收到数据：', end='')
    #         print(res.decode())

    def input_pincode(self, dev_hdl):
        # CALLBACK = CFUNCTYPE(None, c_int)
        # Stop_Search_Device = CALLBACK(self._callback)
        # self.sdk_dll_liu.Register_Callback.argtypes = [c_void_p]
        # self.sdk_dll_liu.Register_Callback(Stop_Search_Device)
        self.sdk_dll.SetRemoteDevicePinCode.argtypes = [c_char_p, c_uint16]
        pin = '1234'
        self.sdk_dll.SetRemoteDevicePinCode(bytes(pin, 'utf-8'), c_uint16(len(pin)))
        self.sdk_dll.BluetPairDevice(dev_hdl)

    def IsConnCallBack(self, datatype, data):
        pass
        # self.isconnflag = 1
        #
        # print(datatype)
        # if datatype == 2:
        #     self.conn_num += 1
        # if datatype == 3:
        #     self.conn_num -= 1
        #
        # # print('当前连接手机：' + str(self.conn_num))
        # C_16DATA = cast(data, POINTER(c_uint16))
        # print(C_16DATA.contents)
        # print(data)
        # print(string_at(data, -1).decode())

    def IsConn(self):
        CALLBACK = CFUNCTYPE(None, c_uint32, c_void_p)
        Stop_Search_Device = CALLBACK(self.IsConnCallBack)
        # self.pDll_su.ATTestSetCallback.argtypes = [CALLBACK]
        self.pDll_su.ATTestSetCallback(Stop_Search_Device)
        # while True:
        #     if self.isconnflag == 1:
        #         # print('已经找到！')
        #         self.isconnflag = 0
        # self.pDll_su.ATTestSetCallback(Stop_Search_Device)

    def strtohex(self, a):
        '''
        字符串转换16进制
        :param a:
        :return:
        '''
        sum1 = 0
        length = len(a) - 1
        for i in range(len(a)):
            if '0' <= a[i] <= '9':
                sum1 += int(a[i]) * (16 ** (length - i))
            if a[i] == 'A' or a[i] == 'a':
                sum1 += 10 * (16 ** (length - i))
            if a[i] == 'B' or a[i] == 'b':
                sum1 += 11 * (16 ** (length - i))
            if a[i] == 'C' or a[i] == 'c':
                sum1 += 12 * (16 ** (length - i))
            if a[i] == 'D' or a[i] == 'd':
                sum1 += 13 * (16 ** (length - i))
            if a[i] == 'E' or a[i] == 'e':
                sum1 += 14 * (16 ** (length - i))
            if a[i] == 'F' or a[i] == 'f':
                sum1 += 15 * (16 ** (length - i))

        return hex(sum1)

    def str_to_addr(self, addr):
        addr_back = []
        if 14 >= len(addr) >= 12:
            for i in range(12, 0, -2):
                addr_back.append(int(self.strtohex(addr[i - 2:i]), 16))
            return addr_back
        else:
            print('连接设备地址错误！')
            return 0

    def read_excel(self):
        '''
        读取excel里面的数据
        :return:
        '''
        # len_type = len(self.type)
        # self.result_data = []  # 读取的结果 二维数组
        # self.data = pd.read_excel(self.path, sheet_name='SPP')
        # self.data = pd.read_excel(r'C:\Users\barrot\Desktop\8051(A02)\test.xlsx', sheet_name='SPP')
        self.data = None
        self.data = pd.read_excel(self.path, sheet_name=self.sheet_name)
        self.type = self.data.columns.values.tolist()            # 得到第一行的值
        # for i in range(len(self.type)-1, -1, -1):
        #     if 'Unnamed' in self.type[i]:
        #         self.type.remove(self.type[i])
        # print(self.type)
        length = len(self.data[self.type[0]])
        for i in range(length):
            temp = []
            for j in range(len(self.type)):
                data_temp = self.data[self.type[j]][i]
                temp.append(data_temp)
            self.result_data.append(temp)

    def make_data(self, index, data):
        result = []
        for i in range(len(data)):
            temp_s = self.result_data[i][index]
            temp_d = []
            if '\CR' in temp_s:
                temp_s = temp_s.split('\CR')
                for reg in temp_s:
                    reg = reg.upper()  # 全转换为大写字母
                    if 'SOC' in reg:
                        reg = reg.split('SOC')  # 拆分AT指令
                        reg = 'soc' + reg[1]
                        temp_d.append(reg)
                    elif 'soc' in reg:
                        reg = reg.split('soc')
                        reg = 'soc' + reg[1]
                        temp_d.append(reg)
                    elif 'REMOTE' in reg:
                        reg = reg.split('REMOTE')
                        reg = 'remote' + reg[1]
                        temp_d.append(reg)
                    elif 'remote' in reg:
                        reg = reg.split('remote')
                        reg = 'remote' + reg[1]
                        temp_d.append(reg)
                    elif 'LOCAL' in reg:
                        reg = reg.split('LOCAL')
                        reg = 'local' + reg[1]
                        temp_d.append(reg)
                    elif 'local' in reg:
                        reg = reg.split('local')
                        reg = 'local' + reg[1]
                        temp_d.append(reg)
            result.append(temp_d)

        return result

    def del_data(self):
        '''
        处理测试用例的表格数据，得到AT指令
        :return:
        '''
        self.read_excel()
        index = self.type.index(self.ini)
        # result_order = []  # 得到初始化条件的AT\BT指令列表
        index1 = self.type.index(self.process)
        # result_order1 = []  # 得到测试过程的AT\BT指令列表

        result_order = self.make_data(index, self.result_data)
        result_order1 = self.make_data(index1, self.result_data)
        # for i in range(len(self.result_data)):
        #     temp_s = self.result_data[i][index]
        #     temp_d = []
        #     if '\CR' in temp_s:
        #         temp_s = temp_s.split('\CR')
        #         for reg in temp_s:
        #             reg = reg.upper()  # 全转换为大小字母
        #             if 'soc' in reg or 'SOC' in reg:
        #                 reg = reg.split('soc+')  # 拆分AT指令
        #                 reg = 'soc+' + reg[1]
        #                 temp_d.append(reg)
        #             if 'remote' in reg or 'REMOTE' in reg:
        #                 reg = reg.split('remote')
        #                 reg = 'AT+' + reg[1]
        #                 temp_d.append(reg)
        #     result_order.append(temp_d)
        #
        # for i in range(len(self.result_data)):
        #     temp_s = self.result_data[i][index1]
        #     temp_d = []
        #     if '\CR' in temp_s:
        #         temp_s = temp_s.split('\CR')
        #         for reg in temp_s:
        #             reg = reg.upper()
        #             if 'AT+' in reg:
        #                 reg = reg.split('AT+')
        #                 reg = 'AT+' + reg[1]
        #                 temp_d.append(reg)
        #             if 'BT+' in reg:
        #                 reg = reg.split('BT+')
        #                 reg = 'AT+' + reg[1]
        #                 temp_d.append(reg)
        #     result_order1.append(temp_d)

        # print('表格读取完成!')
        # self.log.write('表格读取完成!',self.scr)
        return result_order, result_order1

    def get_port(self):
        '''
        检索串口
        :return
        '''
        self.com_list.clear()
        plist = list(serial.tools.list_ports.comports())
        for por in range(0, len(plist)):
            if str(plist[por]).find('COM') >= 0:
                com = str(plist[por])[0:4]
                self.com_list.append(com)
        if len(self.com_list) == 0:
            # print(time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '没有找到对应模块的串口')
            self.log.write('没有找到对应模块的串口', self.scr)

    def AT_order_Send(self, order):
        '''
        发送AT命令
        :param order
        :return:
        '''
        wait_time = 2
        readValue = ''
        self.ser.write(order.encode())
        while self.ser.in_waiting == 0 and wait_time > 0:
            wait_time -= 0.05
            time.sleep(0.05)
        while self.ser.in_waiting > 0:
            reads = self.ser.read(self.ser.in_waiting)
            readValue = readValue + reads.decode('ascii')
        return readValue

    # def AT_order(self, order, res='ming'):  # 修改后
    #     '''
    #     发送AT命令
    #     :param res:
    #     :param order
    #     :return:
    #     '''
    #     if 'COMMAND' in order and '1' in order:
    #         self.isPDU = 1
    #     if 'COMMAND' in order and '0' in order:
    #         self.isPDU = 0
    #
    #     # self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
    #     # self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
    #     # self.pDll_su.ATCmdTestCreateObject()
    #     # self.pDll_su.ATTestOpenSerial(3, 57600)
    #     # self.pDll_su.ATTestOpenThread(1, 0)
    #
    #     # WriteLength = c_int(10)
    #     # ReadLength = c_int(100)
    #     # pDll.ATTestWriteSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
    #     # pDll.ATTestWriteSerial(bytes("AT+GVER?\r\n\0", 'utf-8'), WriteLength, 2000)
    #     # ReadData = create_string_buffer(100)
    #     # pDll.ATTestReadSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
    #     # pDll.ATTestReadSerial(ReadData, ReadLength, 2000)
    #
    #     if 'SEND' in order and self.isPDU == 1:
    #         ##打开Pdu模式测试线程
    #         # CATCmdTestAPI bool PduModeSendData(UINT16 Opcode, UINT16 Conn_hdl, UINT16 Length, char* InData)
    #         reg = order.split('')
    #         pDll.ATTestOpenThread(1, None)
    #         pDll.PduModeSendData.argtypes = [c_uint16, c_uint16, c_uint16, c_char_p]
    #         pDll.PduModeSendData.restype = c_bool
    #         pDll.ATTestCloseThread(1)
    #         time.sleep(1)
    #         return 'pass'
    #     # self.pDll_su.pDll.ATTestOpenSerial(self.com[-1], self.BAUD)
    #
    #     # 打开Pdu模式测试线程
    #     # pDll.ATTestOpenThread(1,0)
    #     # 打开连续模式测试线程
    #     # pDll.ATTestOpenThread(2,0)
    #     # 打开重复模式测试线程
    #     # pDll.ATTestOpenThread(3,10)
    #     # 打开随机模式测试线程
    #     # pDll.ATTestOpenThread(4,0)
    #     self.pDll_su.ATCmdSingleCmdTestRetRes.argtypes = [c_char_p, c_uint16, c_char_p, c_char_p]
    #     self.pDll_su.ATCmdSingleCmdTestRetRes.restype = c_bool
    #
    #     ReadData = create_string_buffer(1000)
    #     s = self.pDll_su.ATCmdSingleCmdTestRetRes(bytes(order, 'utf-8'), len(order), bytes(res, 'utf-8'), ReadData)
    #     print('返回的结果：' + str(s))
    #
    #     self.pDll_su.ATTestCloseSerial()
    #     self.pDll_su.ATCmdTestDeleteObject()
    #
    #     print(ReadData.value.decode())
    #
    #     return ReadData.value.decode()

    def init_spp_service(self, count):
        '''
        初始化spp服务，
        :param count:
        :return:
        '''
        self.sdk_dll.BluetRegisterSPPService.argtype = [c_ubyte]
        self.sdk_dll.BluetRegisterSPPService(count)

    def Disconnection(self, conn, del_hdl=1):
        '''
        断开连接
        :param del_hdl:  是否删除已连接的信息
        :param conn:
        :return:
        '''
        self.log.write('断开SPP连接', self.scr)
        flag = self.sdk_dll.BluetDisConnectRemoteDevice(conn)
        if flag != 1:
            return 0
        if del_hdl == 1:
            if len(self.hConn_info) > 0:
                for i in range(len(self.hConn_info) - 1, -1, -1):
                    if self.hConn_info[i]['id'] == self.devices_id:
                        self.hConn_info.remove(self.hConn_info[i])

        self.conn_num_spp -= 1
        self.log.write('当前SPP连接数量：'+str(self.conn_num_spp), self.scr, mode=1)
        return 1

    def Connection(self, braddr, brclass=0, pin='1234', mode=13):
        '''
        连接手机
        :return:
        '''
        # 定义回调函数
        CALLBACK = CFUNCTYPE(None, c_ulong, POINTER(c_ubyte))
        callBackFunc = CALLBACK(self.SearchCompleteCback)
        self.sdk_dll.SearchCompleCallback(callBackFunc)
        self.sdk_dll.BluetSearchRemoteDevice.argtype = [c_ulong, c_char_p]
        self.sdk_dll.BluetSearchRemoteDevice.restype = c_ulong
        brclass = c_ulong(int(brclass))
        addr = self.str_to_addr(braddr)

        self.sdk_dll.BluetSetLocalMode.argtypes = [c_ushort]
        mode1 = c_ushort(mode)
        self.sdk_dll.BluetSetLocalMode(mode1)  # 设置手机蓝牙发现模式

        self.sdk_dll.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
        pin_code = c_char_p(bytes(pin, 'utf-8'))
        pin_code_len = len(pin)
        self.sdk_dll.BluetSetFixedPinCode(pin_code, pin_code_len)

        if addr == 0:
            print('连接设备的蓝牙地址错误！')
            self.log.write('连接设备的蓝牙地址错误', self.scr)
            return 0
        bdaddr = (c_char * 6)(*addr)
        Search_flag = self.sdk_dll.BluetSearchRemoteDevice(brclass, bdaddr)
        print('模块搜索中！')
        # self.log.write('模块搜索中！', self.scr)
        while True:
            if self.search_flag == 1:
                self.search_flag = 0
                # print('已经搜索到经典蓝牙设备')
                self.log.write('已经搜索到经典蓝牙设备', self.scr)
                self.sdk_dll.BluetStopDiscoverDevice()
                # print('停止搜索')
                self.log.write('停止搜索', self.scr)
                self.sdk_dll.BluetConnectRemoteDevice.argtype = [c_ulong]
                self.sdk_dll.BluetConnectRemoteDevice.restype = c_ulong
                print('设备句柄：' + str(self.dev_hdl))
                self.input_pincode(self.dev_hdl)  # 输入pincode
                conn_hdl = self.sdk_dll.BluetConnectRemoteDevice(self.dev_hdl)
                self.hConn = conn_hdl
                print('连接句柄')
                print(conn_hdl)
                if conn_hdl == 13 or conn_hdl == 0:
                    print(conn_hdl)
                    # print('模块SPP连接失败！')
                    self.log.write('模块SPP连接失败', self.scr)
                    return 0
                else:
                    dev_dic = {}
                    dev_dic['id'] = self.devices_id
                    dev_dic['dev_hdl'] = self.dev_hdl
                    dev_dic['conn_hdl'] = conn_hdl
                    dev_dic['conn_bd_spp'] = braddr
                    dev_dic['spp'] = 'SPP'
                    self.hConn_info.append(dev_dic)
                    # print('模块SPP连接成功！')
                    self.log.write('模块SPP连接成功', self.scr)
                    # self.salve_read_data()                                      # 是否连接成功
                    self.conn_num_spp += 1
                    print('当前SPP连接数量：', end='')
                    # print(self.conn_num_spp)
                    self.log.write('当前SPP连接数量'+str(self.conn_num_spp), self.scr)
                self.conns.append(self.hConn)
                return 1

    def Get_Gatt_Services(self):
        self.sdk_dll.GATT_GetServices.restype = POINTER(BtSdkGATTService)

        self.sdk_dll.ATT_GETUUIDType.argtypes = [c_ushort]
        self.sdk_dll.ATT_GETUUIDType.restype = c_char_p

        self.sdk_dll.GATT_GetCharacteristics.argtypes = [c_ubyte]  # 得到特征 get characteristics
        self.sdk_dll.GATT_GetCharacteristics.restype = POINTER(BtSdkGATTCharacteristic)

        self.sdk_dll.GATT_GetCharacteristicsValue.argtypes = [c_ubyte]  # 得到特征值 get characteristics_value
        self.sdk_dll.GATT_GetCharacteristicsValue.restype = POINTER(BtsdsGATTCharacteristicValueStru)

        self.sdk_dll.GATT_SetCharacteristicsValue.argtypes = [c_char, c_char]  # 设置特征值 set characteristics_value
        self.sdk_dll.GATT_SetCharacteristicsValue.retypes = c_long

        self.sdk_dll.GATT_GetDescriptors.argtypes = [c_ubyte]  # 得到描述 get Descriptors
        self.sdk_dll.GATT_GetDescriptors.restype = POINTER(BtSdkGATTDescriptor)

        self.sdk_dll.GATT_GetDescriptorValue.argtypes = [c_ubyte]  # 得到描述值 get Descriptors_value
        self.sdk_dll.GATT_GetDescriptorValue.restype = POINTER(BtsdkGATTDscriptorValueStru)

        self.sdk_dll.GATT_SetDescriptorValue.argtypes = [c_ubyte, c_ubyte]  # 设置描述值 set Descriptors_value

        try:
            gatt_service = self.sdk_dll.GATT_GetServices()  # 得到服务 get services
            if gatt_service is None:
                print('没有发现服务')
                self.log.write('没有发现服务', self.scr)
                return '没有发现服务'
            else:
                for i in range(int(gatt_service.contents.num)):
                    uuid = gatt_service.contents.service[i].ServiceUuid.ShortUuid
                    src = self.sdk_dll.ATT_GETUUIDType(uuid)
                    print("service id:%d  uuid = %d %s handle = 0x%02x" % (
                        i + 1, uuid, src.decode(), gatt_service.contents.service[i].AttributeHandle))
                    # self.log.write('service id:'+str(i+1)+'uuid = ' + uuid + ' ' + src.decode() + 'handle = ' + gatt_service.contents.service[i].AttributeHandle, self.scr)
                    self.log.write("service id:%d  uuid = %d %s handle = 0x%02x" % (
                        i + 1, uuid, src.decode(), gatt_service.contents.service[i].AttributeHandle), self.scr)
                    gatt_character = self.sdk_dll.GATT_GetCharacteristics(c_ubyte(i + 1))
                    if gatt_character is not None:
                        for j in range(gatt_character.contents.num):
                            uuid_c = gatt_character.contents.character[j].CharacteristicUuid.ShortUuid
                            src_c = self.sdk_dll.ATT_GETUUIDType(uuid_c)
                            self.log.write((" character id:%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
                                j + 1, gatt_character.contents.character[j].AttributeHandle,
                                gatt_character.contents.character[j].CharacteristicValueHandle, src_c.decode())), self.scr)
                            print(" character id:%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
                                j + 1, gatt_character.contents.character[j].AttributeHandle,
                                gatt_character.contents.character[j].CharacteristicValueHandle, src_c.decode()))

                            self.log.write("  [B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
                                gatt_character.contents.character[j].IsBroadcastable,
                                gatt_character.contents.character[j].IsReadable,
                                gatt_character.contents.character[j].IsWritable,
                                gatt_character.contents.character[j].IsSignedWritable,
                                gatt_character.contents.character[j].IsWritableWithoutRespon,
                                gatt_character.contents.character[j].IsIndicatable,
                                gatt_character.contents.character[j].IsNotifiable,
                                gatt_character.contents.character[j].HasExtendedProperties), self.scr)
                            print("  [B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
                                gatt_character.contents.character[j].IsBroadcastable,
                                gatt_character.contents.character[j].IsReadable,
                                gatt_character.contents.character[j].IsWritable,
                                gatt_character.contents.character[j].IsSignedWritable,
                                gatt_character.contents.character[j].IsWritableWithoutRespon,
                                gatt_character.contents.character[j].IsIndicatable,
                                gatt_character.contents.character[j].IsNotifiable,
                                gatt_character.contents.character[j].HasExtendedProperties))

                            gatt_character_val = self.sdk_dll.GATT_GetCharacteristicsValue(c_ubyte(j + 1))
                            if gatt_character_val is not None and gatt_character.contents.character[
                                j].IsReadable == 1:
                                for s in range(gatt_character_val.contents.DataSize):
                                    self.log.write("   value:0x%02x" % gatt_character_val.contents.Data[s], self.scr)
                                    print("   value:0x%02x" % gatt_character_val.contents.Data[s], end='')
                                print('')

                            gatt_des = self.sdk_dll.GATT_GetDescriptors(c_ubyte(j + 1))
                            try:
                                if gatt_des is not None:
                                    for n in range(gatt_des.contents.num):
                                        uuid = gatt_des.contents.des[i].DescriptorUuid.ShortUuid
                                        src = self.sdk_dll.ATT_GETUUIDType(uuid)
                                        self.log.write("    Descriptor id:%d,[Hdl,D:0x%04x] %s" % (
                                            n, gatt_des.contents.des[i].AttributeHandle, src.decode()),self.scr)
                                        print("    Descriptor id:%d,[Hdl,D:0x%04x] %s" % (
                                            n, gatt_des.contents.des[i].AttributeHandle, src.decode()))
                                        self.log.write("    character handle =" +
                                              gatt_des.contents.des[i].CharacteristicHandle, self.scr)
                                        print("    character handle =" +
                                              gatt_des.contents.des[i].CharacteristicHandle)

                                        gatt_des_val = self.sdk_dll.GATT_GetDescriptorValue(c_ubyte(n + 1))
                                        if gatt_des_val is not None:
                                            # print('             gatt_des_val')
                                            if gatt_des_val.contents.DescriptorType == 2:
                                                print("    Notify:%d,Ind:%d" % (
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToNotification,
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToIndication))
                            except Exception as e:
                                print('', end='')
            return 'pass'
        except Exception as e:
            print(e)
            return 'Fail'

    def Gatt_Disconnection(self, handle):
        self.log.write('断开GATT连接', self.scr)
        self.sdk_dll.GATT_EndSession(handle)
        self.conn_num_gatt -= 1
        # print('当前GATT连接数量：', end='')
        # print(self.conn_num_gatt)
        self.log.write('当前GATT连接数量'+str(self.conn_num_gatt), self.scr, mode=1)
        self.sdk_dll.GATT_DeInit()            #清理缓存

    def Gatt_Connection(self, braddr, brclass=0, pin='1234', mode=13):
        self.sdk_dll.GATT_Initial()
        addr = self.str_to_addr(braddr)
        CALLBACK = CFUNCTYPE(None, c_ulong, POINTER(c_ubyte))
        callBackFunc = CALLBACK(self.SearchCompleteCback)
        self.sdk_dll.SearchCompleCallback(callBackFunc)
        self.sdk_dll.BluetSearchRemoteDevice.argtype = [c_ulong, c_char_p]
        self.sdk_dll.BluetSearchRemoteDevice.restype = c_ulong
        brclass = c_ulong(int(brclass))

        self.sdk_dll.BluetSetLocalMode.argtypes = [c_ushort]
        mode = c_ushort(mode)
        self.sdk_dll.BluetSetLocalMode(mode)  # 设置手机蓝牙发现模式

        self.sdk_dll.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
        pin_code = c_char_p(bytes(pin, 'utf-8'))
        pin_code_len = len(pin)
        self.sdk_dll.BluetSetFixedPinCode(pin_code, pin_code_len)
        if addr == 0:
            print('连接设备的蓝牙地址错误！')
            return 0

        bdaddr = (c_char * 6)(*addr)
        Search_flag = self.sdk_dll.BluetSearchRemoteDevice(brclass, bdaddr)
        print('模块搜索中！')
        while True:
            if self.search_flag == 1:
                self.search_flag = 0
                print('已经搜索到BLE设备')
                self.sdk_dll.BluetStopDiscoverDevice()
                print('停止搜索')
                print(self.dev_hdl)
                self.sdk_dll.GATT_Connect.argtypes = [c_ulong]
                flag = self.sdk_dll.GATT_Connect(self.dev_hdl)
                if flag == 0:
                    dev_dic = {}
                    dev_dic['id'] = self.devices_id
                    dev_dic['dev_hdl'] = self.dev_hdl
                    dev_dic['conn_bd_gatt'] = braddr
                    dev_dic['gatt'] = 'GATT'
                    self.hConn_info.append(dev_dic)
                    print('模块GATT连接成功！')
                    self.conn_num_gatt += 1
                    print('当前GATT连接数量：', end='')
                    print(self.conn_num_gatt)
                    return 1
                else:
                    print('模块GATT连接失败')
                    return 0

    def simple_conn(self, con_handle, times):
        '''
        测试多连接
        :param con_handle:
        :param times:
        :return:
        '''
        # print(type(times))
        conn_success_flag = 0
        for i in range(1, times + 1):
            time.sleep(2)
            conn = self.sdk_dll.BluetConnectRemoteDevice(con_handle)
            if conn == 13 or conn == 0:
                print('第' + str(i + 1) + '次连接失败！')
                self.log.write('第' + str(i + 1) + '次连接失败！', self.scr)
            else:
                self.conn_num_spp += 1
                conn_success_flag += 1
                print('第' + str(i + 1) + '次连接成功！')
                self.log.write('第' + str(i + 1) + '次连接成功！', self.scr)
                time.sleep(1)
                self.Disconnection(conn)
                time.sleep(1)
        return conn_success_flag

    def gatt_simple_conn(self, dev_handle, times):
        conn_success_flag = 0
        for i in range(1, times + 1):
            self.sdk_dll.GATT_Initial()
            time.sleep(2)
            conn = self.sdk_dll.GATT_Connect(dev_handle)
            if conn == 1:
                print('第' + str(i + 1) + '次连接失败！')
                self.log.write('第' + str(i + 1) + '次连接失败！', self.scr)
            else:
                self.conn_num_gatt += 1
                conn_success_flag += 1
                print('第' + str(i + 1) + '次连接成功！')
                self.log.write('第' + str(i + 1) + '次连接成功！', self.scr)
                time.sleep(1)
                self.Gatt_Disconnection(dev_handle)
                time.sleep(1)
        return conn_success_flag

    def master_read_data(self):  # 手机收
        self.sdk_dll.BluetGetRemoteBuff.restype = c_char_p
        res = self.sdk_dll.BluetGetRemoteBuff()
        # print('手机接收到数据：', end='')
        # print(res.decode('utf-8', 'ignore'))
        return res.decode('utf-8', 'ignore')

    def salve_read_data(self, length=200, timeout=2000):  # 模块收
        ReadLength = c_int(length)
        ReadData1 = create_string_buffer(length)
        self.pDll_su.ATTestReadSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        self.pDll_su.ATTestReadSerial.restype = c_bool
        read_flag = self.pDll_su.ATTestReadSerial(ReadData1, ReadLength, timeout)
        if read_flag:
            data = ReadData1.raw.decode('utf-8', 'ignore')
            result = ''.join(re.findall(r'[A-Za-z0-9_:]|[\s]|[\n]', data))
            return result.replace('\r', '').replace('\n', ' ')
        else:
            print('没有读到字符串！')
            return 0

    def master_write_data(self, hConn, data, length):  # 手机发
        self.sdk_dll.BluetWriteToRemoteDevice.argtype = [c_ulong, c_char_p, c_ushort]
        self.sdk_dll.BluetWriteToRemoteDevice(hConn, data, length)

    def salve_write_data(self, data, timeout=2000):  # 模块发
        WriteLength = c_int(len(data))
        # WriteLength = c_long(len(data))
        self.pDll_su.ATTestWriteSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        self.pDll_su.ATTestWriteSerial(bytes(data, 'utf-8'), WriteLength, timeout)

    def BT_order_Send(self, order):
        '''
        发送BT指令
        :param order:
        :return:
        '''
        if '=' in order:
            reg = order.split('=')
            if 'NAME' in reg[0]:
                print('设置手机NAME：', end='')
                print(reg[1])
                name = c_char_p(bytes(reg[1], 'utf-8'))
                self.sdk_dll_jiang.BluetSetLocalName.argtype = [c_char_p]
                self.sdk_dll_jiang.BluetSetLocalName(name)
                return 'pass'

            elif 'PIN' in reg[0]:
                self.sdk_dll_jiang.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
                reg[1] = reg[1].replace('\n', '')
                reg[1] = reg[1].replace('\r', '')
                pin_code = c_char_p(bytes(reg[1], 'utf-8'))
                pin_code_len = len(reg[1])
                print('设置手机PIN：', end='')
                print(reg[1])
                self.sdk_dll_jiang.BluetSetFixedPinCode(pin_code, pin_code_len)
                return 'pass'

            elif 'DISCOVER' in reg[0]:
                self.sdk_dll_jiang.BluetSetLocalMode.argtypes = [c_ushort]
                mode = c_ushort(int(reg[1]))
                print('设置手机DISCOVER：', end='')
                print(reg[1])
                self.sdk_dll_jiang.BluetSetLocalMode(mode)
                return 'pass'

            elif 'DISCONNS' in reg[0]:
                if ' ' in reg[1]:
                    res = reg[1].split(' ')
                    if len(res) == 3:
                        print('连断测试开始！')
                        times = int(res[2])
                        self.init_spp_service(1)
                        self.Connection(braddr=res[1], brclass=res[0])
                        self.Disconnection(self.hConn)
                        flag = 0
                        if self.hConn == 0 or self.hConn == 13:
                            print('第1次连接失败！')
                            for i in range(1, int(reg[2])):
                                self.Connection(braddr=res[1], brclass=res[0])
                                self.Disconnection(self.hConn)
                                if self.hConn != 0 and self.hConn != 13:
                                    flag += self.simple_conn(self.dev_hdl, times - i)
                                    break
                                else:
                                    print('第' + str(i) + '次连接失败！')
                        else:
                            flag = 1
                            print('第1次连接成功！')
                            time.sleep(1)
                            print(self.dev_hdl)
                            flag += self.simple_conn(self.dev_hdl, times)

                        if flag == times:
                            return 'pass'
                        else:
                            return 'Fail'

                return 'Fail'

            elif 'CONNS' in reg[0]:
                if ',' in reg[1]:
                    res = reg[1].split(',')
                    count = len(res)
                    for i in res:
                        temp = i.split(' ')
                        self.init_spp_service(count)
                        self.Connection(temp[0], temp[1])

            elif 'CONN' in reg[0]:
                if ' ' in reg[1]:
                    res = reg[1].split(' ')
                    self.init_spp_service(1)
                    print('蓝牙地址长度： ', end='')
                    print(len(res[1]))
                    self.Connection(braddr=res[1], brclass=res[0])
                    if self.hConn == 0 or self.hConn == 13:
                        return 'Fail'
                    return 'pass'
                return 'Fail'

            elif 'SENDS' in reg[0]:
                self.sdk_dll_jiang.BluetWriteToRemoteDevice.argtype = [c_ulong, c_char_p, c_ushort]
                str_len = c_ushort(len(reg[1]))
                str_p = c_char_p(bytes(reg[1], 'utf-8'))
                if len(self.conns) > 0:
                    for conn in self.conns:
                        self.master_write_data(conn, str_len, str_p)
                        self.salve_read_data()
                    for conn in self.conns:
                        self.Disconnection(conn)
                else:
                    print('当前连接数为：1')

            elif 'SEND' in reg[0]:
                if self.hConn == 13 or self.hConn == 0:
                    print('设备未连接！不能发送数据！')
                    return 0
                print('连接句柄：' + str(self.hConn))
                print('手机发送数据：', end='')
                print(reg[1])
                temp = reg[1].split(' ')
                rdata = ''  # 收到的数据
                # 判断是否为循环发送
                if temp[-1].isdigit() and len(temp) > 1:
                    times = int(temp[-1])
                    print('循环发送数据' + str(times) + '次')
                    data = ''  # 发送的数据
                    for i in range(len(temp) - 1):
                        data += temp[i] + ' '
                    data = data[:-1]
                    print('发送的字节长度：' + str(len(data)))
                    str_len = c_ushort(len(data))
                    str_p = c_char_p(bytes(data, 'utf-8'))
                    for i in range(times):
                        self.master_write_data(self.hConn, str_p, str_len)
                        rdata += self.salve_read_data()
                    print(len(data) * times)
                    print(len(rdata))
                    if len(data) * times == len(rdata):
                        print('数据没有丢失')
                        return 'pass'

                    if len(data) * times < len(rdata):
                        print('多收字节数:' + str(len(rdata) - len(data) * times))
                        return '+' + str(len(rdata) - len(data) * times)

                    if len(data) * times > len(rdata):
                        print('丢失字节数' + str(len(data) * times - len(rdata)))
                        return '+' + str(len(data) * times - len(rdata))
                else:
                    print('发送一次')
                    str_len = c_ushort(len(reg[1]))
                    str_p = c_char_p(bytes(reg[1], 'utf-8'))
                    self.master_write_data(self.hConn, str_p, str_len)
                    rdata = self.salve_read_data()
                    if reg[1] in rdata:
                        print('数据完整')
                        return 'pass'
                    else:
                        print('数据丢失')
                        return 'Fail'

            else:
                print('未知的BT指令')
                return 'Fail'

        elif '?' in order:
            if 'BRINFO' in order:
                print('本机信息：')
                self.sdk_dll_jiang.BluetGetLocalInfo.restype = LocalInfo
                self.local_info = self.sdk_dll_jiang.BluetGetLocalInfo()
                print('设备名称：' + str(self.local_info.name.decode('utf-8')))
                print('设备类型：' + str(self.local_info.device_class))
                addr = ''
                for i in range(len(self.local_info.bd_addr)):
                    if len(hex(self.local_info.bd_addr[i])) == 3:
                        addr += '0' + hex(self.local_info.bd_addr[i])[-1]
                    if len(hex(self.local_info.bd_addr[i])) == 4:
                        addr += hex(self.local_info.bd_addr[i])[-2:]
                print('模块地址：' + addr)
                print('设备发现模式：' + str(self.local_info.discover_mode))
                return 'pass'

            elif 'BRADDR' in order:
                print('查询本机蓝牙地址：')
                self.sdk_dll_jiang.BluetGetLocalInfo.restype = LocalInfo
                self.br_name = self.sdk_dll_jiang.BluetGetLocalInfo().decode('utf-8')
                addr = ''
                for i in range(len(self.local_info.bd_addr)):
                    if len(hex(self.local_info.bd_addr[i])) == 3:
                        addr += '0' + hex(self.local_info.bd_addr[i])[-1]
                    if len(hex(self.local_info.bd_addr[i])) == 4:
                        addr += hex(self.local_info.bd_addr[i])[-2:]
                print('模块蓝牙地址：' + addr)
                return 'pass'

            elif 'NAME' in order:
                print('查询本机名字：')
                self.sdk_dll_jiang.BluetGetLocalInfo.restype = LocalInfo
                self.br_name = self.sdk_dll_jiang.BluetGetLocalInfo().decode('utf-8')
                print('设备名称：' + str(self.local_info.name))
                return 'pass'

            elif 'DISCOVERMODE' in order:
                print('查询本机发现模式：')
                self.sdk_dll_jiang.BluetGetLocalInfo.restype = LocalInfo
                self.dis_mode = self.sdk_dll_jiang.BluetGetLocalInfo()
                print('设备发现模式：' + str(self.dis_mode))
                return 'pass'

            elif 'CONNS' in order:
                print('多连接模块信息：')
                self.sdk_dll_jiang.BluetGetCurrConnecInfo.restype = c_ushort
                cur_conn_num = self.sdk_dll_jiang.BluetGetCurrConnecInfo()
                self.sdk_dll_jiang.BluetGetConncetListInfoPointer.restype = POINTER(SimpStruct)

                if cur_conn_num > 0:
                    self.dev_list = self.sdk_dll_jiang.BluetGetConncetListInfoPointer()
                    for i in range(cur_conn_num):
                        print(self.dev_list[i].cls)
                        print(self.dev_list[i].dev_hdl)
                        print(self.dev_list[i].hConn)
                        print(self.dev_list[i].name.decode())
                        print(self.dev_list[i].placement.decode())

                else:
                    print('未连接设备！')

            else:
                print('未知的BT指令')
                return 'Fail'

    def order_analysis(self, order_list):
        '''
        指令解析
        :param order_list:
        :return: dict
        '''
        od_list = []
        for i in range(len(order_list)):
            org = order_list[i].split(' ')
            cmd_dic = {'type': org[0]}
            for reg in range(1, len(org), 2):
                # print(org[reg])
                if len(org[reg]) == 1:
                    cmd_dic[org[reg]] = org[reg + 1]
                else:
                    cmd_dic[org[reg][1:]] = org[reg + 1]
            od_list.append(cmd_dic)
        return od_list

    def soc_del(self, order, at_cmd=None, data=None, dt=None, dl=None):  # AT指令
        '''
        处理soc命令
        :param order: soc命令
        :param at_cmd: AT指令
        :param data: 发送的数据
        :param dt: 发送数据的类型：int\string
        :param dl: 发送数据的长度
        :return:
        '''
        MTU = 200
        if at_cmd is not None:
            print('发送AT指令：' + at_cmd)
            self.salve_read_data()  # 清空模块缓存
            self.salve_write_data(at_cmd + '\r')
            # time.sleep(0.1)
            read = self.salve_read_data()
            if read == 0:
                return '模块没有响应'
            print("读取：" + str(read))
            return str(read)
        else:
            if data is not None:
                print('发送数据：' + data)
                self.salve_write_data(data)
                time.sleep(0.1)
                read_data = self.master_read_data()
                if read_data == 'no str':
                    return '手机未收到数据'
                else:
                    print(len(read_data))
                    print('手机收到数据：' + read_data)
                    return read_data
            #     ***********************************************
            else:
                if dt is not None and dl is not None:
                    read_data = ' '
                    read_length = 0
                    if dt == 'STRING' or dt == 'string':
                        print('模块发送数据: ' + str(dl) + '位')
                        data = ''.center(MTU, 's')
                        length = int(int(dl) / MTU)
                        leave = int(dl) % MTU
                        for i in range(length):
                            self.salve_write_data(data)
                            time.sleep(0.1)
                            read_data = self.master_read_data()
                            if 'no str' not in read_data:
                                read_length += len(read_data)
                        self.salve_write_data(''.center(leave, 's'))
                        time.sleep(0.1)
                        read_data = self.master_read_data()
                        # print('手机收到数据：' + read_data)
                        if 'no str' not in read_data:
                            read_length += len(read_data)
                        if read_length == 0:
                            return '手机未收到数据!'
                        if read_length == int(dl):
                            return 'pass'
                        elif read_length > int(dl):
                            return '数据增加' + str(read_length - int(dl))
                        elif read_length < int(dl):
                            return '数据丢失' + str(int(dl) - read_length)
                    if dt == 'INT' or dt == 'int':
                        data = ''.center(MTU, '1')
                        length = int(int(dl) / MTU)
                        leave = int(dl) % MTU
                        for i in range(length):
                            self.salve_write_data(data)
                            read_data += self.master_read_data()
                        self.salve_write_data(''.center(leave, 's'))
                        read_data += self.master_read_data()
                        if 'no str' in read_data:
                            print('手机没有读到数据')
                            return '手机收到数据'
                        if len(read_data) - 1 == int(dl):
                            return 'pass'
                        elif len(read_data) - 1 > int(dl):
                            return '数据增加' + str(len(read_data) - 1 - int(dl))
                        elif len(read_data) - 1 < int(dl):
                            return '数据丢失' + str(int(dl) - int(len(read_data) - 1))
                else:
                    return 'soc指令错误'
        return 'soc指令错误'

    def remote_del(self, order, device_id='0', cmd=None, bd=None, bt='public', t=None, dl=None, cf=None):
        '''
        处理remote命令
        :param cf:
        :param dl:
        :param order: remote命令
        :param device_id: remote id
        :param cmd: 执行命令
        :param bd: 蓝牙地址
        :param bt: 蓝牙地址类型
        :param t: 蓝牙连接类型
        :return: pass\Fail
        '''
        if cmd == 'inquiry' or cmd == 'INQUIRY':
            if t == 'GATT' or t == 'GATT':
                flag = self.Get_Gatt_Services()
                return flag

            print('本机信息：')
            self.sdk_dll.BluetGetLocalInfo.restype = LocalInfo
            self.local_info = self.sdk_dll.BluetGetLocalInfo()
            print('设备名称：' + str(self.local_info.name.decode('utf-8')))
            print('设备类型：' + str(self.local_info.device_class))
            addr = ''
            for i in range(len(self.local_info.bd_addr)):
                if len(hex(self.local_info.bd_addr[i])) == 3:
                    addr += '0' + hex(self.local_info.bd_addr[i])[-1]
                if len(hex(self.local_info.bd_addr[i])) == 4:
                    addr += hex(self.local_info.bd_addr[i])[-2:]
            print('模块地址：' + addr)
            print('设备发现模式：' + str(self.local_info.discover_mode))
            return 'pass'

        elif cmd == 'CONN' or cmd == 'conn':
            # print('1')
            # print(self.data['级别'])
            if bd is not None:
                if bt == 'public' or bt == 'PUBLIC':
                    print('地址类型：public ')
                elif bt == 'random' or bt == 'RANDOM':
                    print('地址类型：random')
                if t == 'SPP' or t == 'spp':
                    print('进行spp连接')
                    self.log.write('进行spp连接', self.scr)
                    # print('2')
                    # print(self.data['级别'])
                    self.init_spp_service(1)
                    flag = self.Connection(braddr=bd)
                    # print('3')
                    # print(self.data['级别'])
                    if flag:
                        return 'pass'
                    return 'spp连接失败'
                if t == 'GATT' or t == 'gatt':
                    print('进行GATT连接')
                    self.log.write('进行GATT连接', self.scr)
                    flag = self.Gatt_Connection(braddr=bd)
                    if flag:
                        return 'pass'
                    return 'gatt连接失败'
            else:
                print('蓝牙地址为空')
                return '蓝牙地址为空'

        elif cmd == 'disc' or cmd == 'DISC':
            if t == 'SPP' or t == 'spp':
                print('断开SPP连接')
                # self.log.write('断开SPP连接', self.scr)
                if len(self.hConn_info) > int(device_id):
                    if 'conn_hdl' in self.hConn_info[int(device_id)].keys():
                        self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'])
                        return 'pass'
                return '未建立SPP连接'
            if t == 'GATT' or t == 'gatt':
                print('断开GATT连接')
                self.log.write('断开GATT连接', self.scr)
                if len(self.hConn_info) > int(device_id):
                    if 'dev_hdl' in self.hConn_info[int(device_id)]:
                        self.Gatt_Disconnection(self.hConn_info[int(device_id)]['dev_hdl'])
                        return 'pass'
                return '未建立GATT连接'

        elif cmd == 'send' or cmd == 'SEND':
            if int(dl) > 0:
                MTU = int(dl)
                data = ''.center(MTU, 's')
                self.salve_read_data()  # 清空模块的缓存区
                read_data = ' '
                length = int(int(dl) / MTU)
                leave = int(dl) % MTU
                # try:
                if len(self.hConn_info) > int(device_id):
                    if 'conn_hdl' in self.hConn_info[int(device_id)].keys():
                        for i in range(length):
                            str_len = c_ushort(len(data))
                            str_p = c_char_p(bytes(data, 'utf-8'))
                            # print(self.hConn_info[int(device_id)]['conn_hdl'])
                            self.master_write_data(self.hConn_info[int(device_id)]['conn_hdl'], str_p, str_len)
                            read_data += self.salve_read_data()
                    else:
                        return '未建立连接'
                else:
                    return '未建立连接'
                # except Exception as s:
                #     print(s)
                #     print('未建立连接')
                #     return '未建立连接'
                data = ''.center(leave, 's')
                str_len = c_ushort(len(data))
                str_p = c_char_p(bytes(data, 'utf-8'))
                self.master_write_data(self.hConn_info[int(device_id)]['conn_hdl'], str_p, str_len)
                read_data += self.salve_read_data()
                print('模块收到数据：' + str(len(read_data) - 1))
                if len(read_data) == 0:
                    return '模块没有收到数据'
                self.log.write('模块收到数据长度' + str(len(read_data) - 1), self.scr)
                if len(read_data) - 1 == int(dl):
                    return 'pass'
                elif len(read_data) - 1 > int(dl):
                    return '数据增加' + str(len(read_data) - 1 - int(dl))
                elif len(read_data) - 1 < int(dl):
                    return '数据丢失' + str(int(dl) - int(len(read_data) - 1))
            else:
                return 'SEND指令错误'

        elif cmd == 'recv' or cmd == 'RECV':
            time.sleep(0.1)
            master_read = self.master_read_data()
            if len(master_read) > 0:
                return 'pass'
            return '未收到数据'

        elif cmd == 'DISCS' or cmd == 'discs':
            if int(cf) > 0:
                times = int(cf)  # 默认10次
            else:
                return 'DISCS指令错误'
            success_times = 0
            if t == 'SPP' or t == 'spp':
                self.init_spp_service(1)
                flag = self.Connection(braddr=bd)
                if flag:
                    # print('第1次连接成功')
                    self.log.write('第1次连接成功', self.scr)
                    self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'], del_hdl=0)
                    success_times = 1
                    success_times += self.simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - flag)
                else:
                    # print('第1次连接失败')
                    self.log.write('第1次连接失败', self.scr)
                    for i in range(1, times):
                        flag = self.Connection(braddr=bd)
                        if flag:
                            self.log.write('第' + str(i + 1) + '次连接成功', self.scr)
                            self.Disconnection(self.hConn_info[int(device_id)]['conn_hdl'])
                            success_times = 1
                            success_times += self.simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - i)
                        else:
                            self.log.write('第' + str(i + 1) + '次连接失败', self.scr)
                # print(success_times)
                if success_times == times:
                    return 'pass'
                return '连接'+str(times)+'次' + '成功' + str(success_times)+'次'
            if t == 'GATT' or t == 'gatt':
                flag = self.Gatt_Connection(braddr=bd)
                if flag:
                    self.log.write('第1次连接成功', self.scr)
                    self.Gatt_Disconnection(self.hConn_info[int(device_id)]['dev_hdl'])
                    success_times = flag
                    success_times += self.gatt_simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - flag)
                else:
                    self.log.write('第1次连接失败', self.scr)
                    for i in range(1, times):
                        flag = self.Gatt_Connection(braddr=bd)
                        if flag:
                            self.log.write('第'+str(i+1)+'次连接成功', self.scr)
                            self.Gatt_Disconnection(self.hConn_info[int(device_id)]['dev_hdl'])
                            success_times = 1
                            success_times += self.gatt_simple_conn(self.hConn_info[int(device_id)]['dev_hdl'], times - i)
                        else:
                            self.log.write('第' + str(i + 1) + '次连接失败', self.scr)
                if success_times == times:
                    return 'pass'
                return '连接' + str(times) + '次' + '成功' + str(success_times) + '次'

        else:
            return 'remote指令错误'

    def local_del(self, order, cmd=None, bd=None, bt='public', t=None):  # 待拓展
        '''

        :param order:
        :param cmd:
        :param bd:
        :param bt:
        :param t:
        :return:
        '''

        return 'pass'

    def order_del_list(self, order_list):
        result_list_tmp = []
        # time.sleep(0.1)
        c = None
        d = None
        dt = None
        dl = None
        bd = None
        bt = 'public'
        t = None
        id = '0'
        result = ''
        cf = None
        order_del_list = self.order_analysis(order_list)
        for order in order_del_list:
            print(order)
            if order['type'] == 'soc' or order['type'] == 'SOC':
                if 'C' in order.keys():
                    c = order['C']
                if 'D' in order.keys():
                    d = order['D']
                if 'DT' in order.keys():
                    dt = order['DT']
                if 'DL' in order.keys():
                    dl = order['DL']
                if 'c' in order.keys():
                    c = order['c']
                if 'd' in order.keys():
                    d = order['d']
                if 'dt' in order.keys():
                    dt = order['dt']
                if 'dl' in order.keys():
                    dl = order['dl']
                result = self.soc_del(order, at_cmd=c, data=d, dt=dt, dl=dl)
                c = None
                d = None
                dt = None
                dl = None
                if result is not None and result != ' ':
                    self.log.write(result, self.scr)
            elif order['type'] == 'remote' or order['type'] == 'REMOTE':
                if 'I' in order.keys():
                    id = order['I']
                if 'C' in order.keys():
                    c = order['C']
                if 'BD' in order.keys():
                    bd = order['BD']
                if 'BT' in order.keys():
                    bt = order['BT']
                if 'T' in order.keys():
                    t = order['T']
                if 'DL' in order.keys():
                    dl = order['DL']
                if 'CF' in order.keys():
                    cf = order['CF']
                if 'i' in order.keys():
                    id = order['i']
                if 'c' in order.keys():
                    c = order['c']
                if 'bd' in order.keys():
                    bd = order['bd']
                if 'bt' in order.keys():
                    bt = order['bt']
                if 't' in order.keys():
                    t = order['t']
                if 'dl' in order.keys():
                    dl = order['dl']
                if 'cf' in order.keys():
                    cf = order['cf']
                result = self.remote_del(order, device_id=id, cmd=c, bd=bd, bt=bt, t=t, dl=dl, cf=cf)
                c = None
                cf = None
                dl = None
                bd = None
                bt = 'public'
                t = None
                id = '0'
            elif order['type'] == 'local' or order['type'] == 'LOCAL':
                if 'C' in order.keys() or 'c' in order.keys():  # inquiry/conn/disc/send/recv
                    c = order['C']
                if 'BD' in order.keys() or 'bd' in order.keys():
                    bd = order['BD']
                if 'BT' in order.keys() or 'bt' in order.keys():
                    bt = order['BT']
                if 'T' in order.keys() or 't' in order.keys():
                    t = order['T']
                result = self.local_del(order, cmd=c, bd=bd, bt=bt, t=t)
                c = None
                bd = None
                bt = 'public'
                t = None
            else:
                print('未知错误！')
                result_list_tmp.append('指令错误！')
            # print(result)
            result_list_tmp.append(result)
        return result_list_tmp

    def cmdtestlist(self, order_list):
        '''
        处理串口命令和串口返回信息
        :param order_list:
        :return:
        '''
        result_list_tmp = []
        time.sleep(0.1)
        for order in order_list:
            if 'AT+' in order:
                # cmdResp = self.AT_order_Send(order + '\r\n')
                cmdResp = self.AT_order((order + '\r\n'))
                if cmdResp == '':
                    result_list_tmp.append('RETURN NULL')  # ##################################
                else:
                    rsp_list = cmdResp.splitlines()
                    for rsp in rsp_list:
                        if rsp != '' and rsp != '+OK':
                            result_list_tmp.append(rsp)
                            continue

            if 'BT+' in order:
                cmdResp = self.BT_order_Send(order)
                result_list_tmp.append(cmdResp)

        return result_list_tmp

    def read_com(self):
        """
        创建读取串口按钮控件
        :return:
        """
        self.read = Button((self.root), text='READ_COM', relief=GROOVE, command=self.thread_choose_com, bg='green',
                           fg='white',
                           height=1,
                           width=12)

    def test_contents(self):
        """
        创建单选控件
        :return:
        """
        self.xVariable = StringVar()
        self.contents = ttk.Combobox((self.root), textvariable=(self.xVariable), state='readonly', width=15)
        self.contents_label = Label(self.root, text="Port:")

    def process_bar(self, percent, start_str='', total_length=0):
        """
        进度条
        :return:   {:0>4.1f}%
        """
        bar = ''.join(["\033[1;31;41m%s\033[0m" % '   '] * int(percent * total_length)) + ''
        bar = '\r' + start_str + bar.ljust(total_length) + ' {:.2f}%'.format(percent * 100)
        print(bar, end='', flush=True)
        self.scr.insert(END, time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '→' ' {:.2f}%'.format(percent * 100) + '\n')
        self.scr.see(END)
        self.scr.update()
        if percent != 1:
            self.scr.delete("%s-1l" % INSERT, INSERT)
        else:
            self.log.write('执行完成!', self.scr)

    def thread_choose_com(self):
        t = threading.Thread(target=self.option_com)
        t.setDaemon(True)     # 同时进行
        t.start()
        # t.join()            # 阻塞调用

    def option_com(self):
        '''
        选择串口
        :return:
        '''
        self.read.config(state='disable')
        self.get_port()
        com_list = tuple(self.com_list)
        self.contents['value'] = com_list
        if len(self.contents['value']) == 0:
            tkinter.messagebox.showinfo(title='Notice', message='没有选中串口\n')
            self.log.write('没有选中串口', self.scr)
        else:
            self.contents.current(0)
        self.read.config(state='normal')

    def selectPath(self):
        '''
        选择路径
        :return:
        '''
        path = askopenfilename(filetypes=[('Python', '*.xlsx *.xls'), ('All files', '*')])
        if len(path) > 0:
            self.test_path_entry.delete(0, END)
            self.test_path_entry.insert(END, path)
            self.log.write('测试用例路径：' + path, self.scr)
        self.path = path

    def get_test_path(self):
        """
        得到测试用例的路径
        :return:
        """
        self.var1 = StringVar()
        self.test_path_entry = Entry(self.root, textvariable=self.var1, bd=2, width=57, state='normal')
        self.test_path_label = Label(self.root, text="测试用例路径:", height=1)
        self.test_path_button = Button(self.root, text="选择文件", command=self.selectPath, height=1)

    def simple_test_map(self):
        var = StringVar()
        self.simple_test_text = Text(self.root, width=109, height=5)
        # self.test_path_label = Label(self.root, text="test case directory:", height=1)
        # self.simple_test_button = Button(self.root, text="发送", command=self.selectPath, height=1)

    def logo(self):
        self.console_label = Label(self.root, text="Console", height=1)
        self.simple_test_logo = Label(self.root, text='测试输入框', height=1)

    def get_baud(self):
        '''
        创建BAUD控件
        :return:
        '''
        self.var2 = StringVar(value='115200')
        self.test_baud_entry = Entry((self.root), textvariable=(self.var2), bd=2, width=12, state='normal')
        self.test_baud_label = Label(self.root, text="BAUD:", height=1)

    def execute_script(self):
        '''
        运行脚本
        :return:
        '''
        self.scr.delete(1.0, END)
        self.execute_button.config(state='disable')
        if len(self.test_baud_entry.get()) == 0:
            tkinter.messagebox.showinfo(title='Notice', message='BAUD不能为空')
            self.log.write('BAUD为空!!', self.scr)
            self.execute_button.config(state='normal')
        else:
            if not self.test_baud_entry.get().isdigit():
                tkinter.messagebox.showinfo(title='Notice', message='BAUD有非法字符!!\n请重新输入!!')
                self.log.write('BAUD有非法字符!!', self.scr)
                self.execute_button.config(state='normal')
            else:
                self.BAUD = int(self.test_baud_entry.get())
                self.log.write('BAUD：' + str(self.BAUD), self.scr)
                if len(self.contents.get()) == 0:
                    tkinter.messagebox.showinfo(title='Notice', message='串口为空!!')
                    self.log.write('串口为空!!', self.scr)
                else:
                    self.com = self.contents.get()
                    if len(self.test_path_entry.get()) > 0:
                        self.path = self.test_path_entry.get()
                        self.log.write(' 测试用例路径: ' + self.path, self.scr)
                        t = threading.Thread(target=self.test_spp)
                        t.setDaemon(True)
                        t.start()

                    elif len(self.simple_test_text.get('1.0', END)) > 0:
                        test = self.simple_test_text.get('1.0', END)
                        test = test.replace('\n', '')
                        self.simple_test(test)
                        self.execute_button.config(state='normal')

                    elif len(self.test_path_entry.get()) <= 0:
                        tkinter.messagebox.showinfo(title='Notice', message='测试用例路径为空！')
                        self.log.write('测试用例路径为空！', self.scr)
                        self.execute_button.config(state='normal')

    def simple_test(self, text):
        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        self.pDll_su.ATTestOpenSerial.retypes = c_bool
        self.pDll_su.ATCmdTestCreateObject()
        CALLBACK = CFUNCTYPE(None, c_uint32, c_void_p)
        Stop_Search_Device = CALLBACK(self.IsConnCallBack)
        self.pDll_su.ATTestSetCallback.argtypes = [CALLBACK]
        self.pDll_su.ATTestSetCallback(Stop_Search_Device)

        open_flag = self.pDll_su.ATTestOpenSerial(int(self.com[-1]), int(self.BAUD))
        if open_flag == 0:
            self.log.write('串口打开失败', self.scr)
            self.pDll_su.ATTestCloseSerial()
            self.pDll_su.ATCmdTestDeleteObject()
            return 0
        # self.pDll_su.ATTestOpenThread(5, 0)

        test = text.replace('\n', '')
        if '\CR' in test:
            order = test.split('\CR')
            print(order)
            self.order_del_list(order)
        else:
            order_list = [test]
            self.order_del_list(order_list)

        # self.pDll_su.ATTestCloseThread(5)
        self.pDll_su.ATTestCloseSerial()
        self.pDll_su.ATCmdTestDeleteObject()

    def execute(self):
        '''
        运行控件
        :return:
        '''
        self.execute_button = Button(self.root, text='EXECUTE', relief=GROOVE, command=self.execute_script,
                                     bg='green', fg='white',
                                     height=1,
                                     width=12)

    def console(self):
        '''
        信息输出框
        :return:
        '''
        self.scr = tkinter.scrolledtext.ScrolledText((self.root), width=95, height=20, font=('楷书', 11))

    def creat_map(self):
        '''
        创建TK界面
        :return:
        '''
        self.read_com()
        self.test_contents()
        self.get_test_path()
        self.get_baud()
        self.execute()
        self.console()
        self.logo()
        self.simple_test_map()
        self.console_label.place(x=380, y=6)
        self.simple_test_logo.place(x=380, y=350)
        self.simple_test_text.place(x=10, y=380)
        self.scr.place(x=10, y=30)
        self.test_path_entry.place(x=190, y=485)
        self.test_path_label.place(x=70, y=482)
        self.test_path_button.place(x=600, y=480)
        self.read.place(x=100, y=575)
        self.execute_button.place(x=600, y=575)
        self.test_baud_entry.place(x=605, y=530)
        self.test_baud_label.place(x=555, y=530)
        self.contents_label.place(x=100, y=530)
        self.contents.place(x=140, y=530)
        self.root.mainloop()

    def test(self):
        '''
        开始执行
        :return:
        '''
        index1 = self.type.index(self.result)
        # index1 = self.type.index(self.process)
        try:
            # self.ser = serial.Serial(self.com, self.BAUD)
            self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
            self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
            self.pDll_su.ATCmdTestCreateObject()
            self.pDll_su.ATTestOpenSerial(int(self.com[-1]), int(self.BAUD))
        except:
            # print(time.strftime('%Y-%m-%d %H:%M:%S') + ' ' + '串口打开失败！')
            tkinter.messagebox.showinfo(title='Notice', message='Failed to open serial port!')
            self.log.write('Failed to open serial port!', self.scr)
            return False
        result_order, result_order1 = self.del_data()
        self.length = len(self.data[self.ini])  # 总长度
        for i in range(0, len(self.data[self.type[0]])):
            ini = self.cmdtestlist(result_order[i])  # 前置条件############################################
            process = self.cmdtestlist(result_order1[i])  # 测试步骤############################################
            regs = ''
            flag = 0
            for reg in ini:
                flag += 1
                regs += str(flag) + '、' + reg + '\n\t'
            for reg in process:
                flag += 1
                regs += str(flag) + '、' + reg + '\n\t'
            if len(process) > 0:
                self.data.loc[i, self.type[index1]] = regs
            if pd.isnull(self.data[self.type[index1]][i]):
                self.data.loc[i, self.type[index1]] = 'Fail'
            # self.process_bar((i + 1) / self.length, start_str='进度', total_length=40)
        print(self.data[self.type[index1]])

    def test_data(self):
        BT_LIST = ['BT+NAME=ming', 'BT+DISCOVER=13', 'BT+PIN=1234', 'BT+CONN=0 EFE18BD6CC1C', 'BT+SEND=1234567',
                   'BT+BRINFO?', 'BT+CONNS?', 'BT+DISCONNS=0 EFE18BD6CC1C 10']

        BT = ['BT+NAME=ming', 'BT+DISCOVER=13', 'BT+PIN=1234', 'BT+CONN=0 488367831500', 'BT+CONNS?', 'BT+SEND'
                                                                                                      '=thisssssss is a string 10']

        BT_TEST = ['BT+NAME=ming', 'BT+DISCOVER=13', 'BT+PIN=1234', 'BT+CONN= 0 488367831500']

        # AT = ['AT+NAME?', 'AT+PIN?', 'AT+GVER?']
        # self.sdk_dll_jiang.BluetFuncInterfaceInit()
        # PduModeGetListData(UINT16 Conn_hdl)

        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        self.pDll_su.ATCmdTestCreateObject()
        self.pDll_su.ATTestOpenSerial(3, 921600)

        # self.pDll_su.ATTestOpenThread(1, 0)

        # self.salve_read_data()
        # self.cmdtestlist(BT)
        # self.AT_order(AT)

        # self.pDll_su.ATTestCloseSerial()
        # self.pDll_su.ATCmdTestDeleteObject()
        # cmd = 'AT+NAME?\r'
        # WriteLength = c_int(len(cmd))
        # print(WriteLength)
        # ReadLength = c_int(100)
        # self.pDll_su.ATTestWriteSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        # self.pDll_su.ATTestWriteSerial(bytes(cmd, 'utf-8'), WriteLength, 2000)
        self.salve_write_data('AT+GVER\r')
        read = self.salve_read_data()
        print(read)
        # ReadData1 = create_string_buffer(100)
        # self.pDll_su.ATTestReadSerial.argtypes = [c_char_p, POINTER(c_int), c_uint32]
        # self.pDll_su.ATTestReadSerial.restype = c_bool
        # s = self.pDll_su.ATTestReadSerial(ReadData1, ReadLength, 2000)
        # e = str(ReadData1.raw.decode())
        # print(s)
        # #
        # print(ReadData1.raw.decode())
        # s = ReadData.value.replace('\r',' ')
        # print(s.encode())

    def test_spp(self):
        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]

        self.pDll_su.ATTestOpenSerial.retypes = c_bool
        self.pDll_su.ATCmdTestCreateObject()

        CALLBACK = CFUNCTYPE(None, c_uint32, c_void_p)
        Stop_Search_Device = CALLBACK(self.IsConnCallBack)
        self.pDll_su.ATTestSetCallback(Stop_Search_Device)

        open_flag = self.pDll_su.ATTestOpenSerial(int(self.com[-1]), int(self.BAUD))


        if open_flag == 0:
            # print('串口打开失败！')
            self.log.write('串口打开失败', self.scr)
            self.pDll_su.ATTestCloseSerial()
            self.pDll_su.ATCmdTestDeleteObject()
            self.execute_button.config(state='normal')
            return 0
        result_order, result_order1 = self.del_data()
        index1 = self.type.index(self.result)              # 写结果的下标
        expecte_index = self.type.index(self.expecte)
        question_index = self.type.index(self.question)
        # result = []
        # expecte = []
        self.length = len(self.data[self.ini])             # 总长度
        for i in range(0, len(self.data[self.type[0]])):
            ini = self.order_del_list(result_order[i])  # 前置条件############################################
            process = self.order_del_list(result_order1[i])  # 测试步骤############################################
            regs = ''
            flag = 0
            success_flag = 1
            print('条件：', end='')
            print(ini)
            for reg in ini:
                if 'pass' == reg:
                    continue
                elif reg is not None:
                    if ' ' in reg:
                        temp = reg.split(' ')
                        print(self.data[self.expecte][i])
                        for temp_reg in temp:
                            if temp_reg in self.data[self.expecte][i]:
                                pass
                            elif temp_reg is None:
                                pass
                            else:
                                flag += 1
                                regs += str(flag) + '、' + temp_reg + '\n\t'
                                success_flag = 0
                    elif reg in self.data[self.expecte][i]:
                        pass
                    else:
                        success_flag = 0
                        flag += 1
                        regs += str(flag) + '、' + reg + '\n\t'
            print('过程：', end='')
            print(process)
            for reg in process:
                if 'pass' == reg:
                    continue
                elif reg is not None:
                    if ' ' in reg:
                        temp = reg.split(' ')
                        for temp_reg in temp:
                            if temp_reg in self.data[self.expecte][i]:
                                print(temp_reg)
                                pass
                            elif temp_reg is None:
                                pass
                            else:
                                flag += 1
                                regs += str(flag) + '、' + temp_reg + '\n\t'
                                success_flag = 0
                    elif reg in self.data[self.expecte][i]:
                        pass
                    else:
                        success_flag = 0
                        flag += 1
                        regs += str(flag) + '、' + reg + '\n\t'

            if success_flag == 1:
                self.data.loc[i, self.type[index1]] = 'Pass'

            if success_flag == 0:
                self.data.loc[i, self.type[index1]] = 'Fail'
                self.data.loc[i, self.type[question_index]] = regs

            if pd.isnull(self.data[self.type[index1]][i]):
                self.data.loc[i, self.type[index1]] = ' '

            if pd.isnull(self.data[self.type[question_index]][i]):
                self.data.loc[i, self.type[question_index]] = ' '

            # self.process_bar((i + 1) / self.length, start_str='进度', total_length=40)
        print(self.data[self.type[index1]])
        print(self.data[self.type[question_index]])

        result = self.data[self.type[index1]].values.tolist()
        question = self.data[self.type[question_index]].values.tolist()

        self.execute_button.config(state='normal')
        self.log.write('用例执行完成！', self.scr)
        # self.pDll_su.ATTestCloseThread(5)
        self.pDll_su.ATTestCloseSerial()
        self.pDll_su.ATCmdTestDeleteObject()
        self.test_excel(result, question)

    def test_gatt(self):
        self.read_excel()
        # s = self.data.columns.values
        # w = self.data['类型'].values.tolist()
        # print(len(w))
        # print(w)
        # gatt = ['remote -C conn -BD C015834563E4 -T GATT', 'remote -C inquiry -T GATT', 'remote -C disc -T GATT']
        # # result = self.order_del_list(gatt)
        # spp = ['remote -C discs -BD 001583678348 -T SPP']
        # # self.order_del_list(spp)
        # self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        # # self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        # self.pDll_su.ATTestOpenSerial.retypes = c_bool
        # self.pDll_su.ATCmdTestCreateObject()
        # open_flag = self.pDll_su.ATTestOpenSerial(3, 921600)
        # if open_flag == 0:
        #     print('打开失败！')
        #     return 0
        # self.Connection('001583678348')
        # # while True:
        # self.salve_write_data('sssssssssss')
        # time.sleep(1)
        # read = self.master_read_data()
        # print(read)

    def spp_t(self):
        # self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        # # self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        # self.pDll_su.ATTestOpenSerial.retypes = c_bool
        # self.pDll_su.ATCmdTestCreateObject()
        # flag = self.pDll_su.ATTestOpenSerial(3, 921600)
        # if flag == 0:
        #     print('串口打开失败!')
        #     return 0
        # self.salve_write_data('AT+GVER' + '\r')
        # time.sleep(0.1)
        # read = self.salve_read_data()
        # print(read)
        # print("读取：" + str(read))
        # return str(read)
        # self.salve_write_data('AT+GVER\r')
        # read = self.salve_read_data()
        # print(read)
        # self.Gatt_Connection('001583678348')
        #  'soc -C AT+GVER', 'soc -C AT+NAME?' 'remote -C conn -BD 001583678348 -T SPP',
        spp = ['remote -C DISCS -BD 001583678348 -T SPP']
        # self.remote_del(order=spp[0], cmd='conn', bd='001583678348', t='SPP')

        self.pDll_su.ATTestOpenSerial.argtypes = [c_uint16, c_uint32]
        self.pDll_su.ATTestOpenThread.argtypes = [c_uint32, c_uint16]
        self.pDll_su.ATCmdTestCreateObject()
        self.pDll_su.ATTestOpenSerial(3, 921600)
        #
        # print(self.order_del_list(spp))
        # bd = '001583678348'
        # bd1 = '001583000013'
        # self.order_del_list(spp)

        # while True:
        #     self.master_read_data()
        # self.order_del_list(spp)

        # self.salve_write_data('asdfasdf')
        # time.sleep(0.01)
        # self.master_read_data()
        #
        # self.salve_write_data('asdfasdfsdfasdf')
        # time.sleep(0.01)
        # self.master_read_data()
        #
        # self.salve_write_data('asdfasdfasdfasdfasdfsdf')
        # time.sleep(0.01)
        # self.master_read_data()
        #
        # self.salve_write_data('asdfasfddddddddddddddddddddasdf')
        # time.sleep(0.01)
        # self.master_read_data()
        #
        # self.salve_write_data('asdssssssssssssssssssssssssssssfasdf')
        # time.sleep(0.1)
        # self.master_read_data()

    def gatt_t(self):
        # self.sdk_dll.BluetFuncInterfaceInit()
        # self.sdk_dll.GATT_Initial()
        bd = '001583678348'
        bd1 = '001583000013'
        bd2 = 'C015834563E4'
        # addr = self.str_to_addr(bd2)
        # CALLBACK = CFUNCTYPE(None, c_ulong, POINTER(c_ubyte))
        # callBackFunc = CALLBACK(self.SearchCompleteCback)
        # self.sdk_dll.SearchCompleCallback(callBackFunc)
        # self.sdk_dll.BluetSearchRemoteDevice.argtype = [c_ulong, c_char_p]
        # self.sdk_dll.BluetSearchRemoteDevice.restype = c_ulong
        # brclass = c_ulong(0)
        # # addr = self.str_to_addr(braddr)
        #
        # self.sdk_dll.BluetSetLocalMode.argtypes = [c_ushort]
        # mode = c_ushort(13)
        # self.sdk_dll.BluetSetLocalMode(mode)  # 设置手机蓝牙发现模式
        #
        # self.sdk_dll.BluetSetFixedPinCode.argtype = [c_char_p, c_ushort]
        # pin = '1234'
        # pin_code = c_char_p(bytes(pin, 'utf-8'))
        # pin_code_len = len(pin)
        # self.sdk_dll.BluetSetFixedPinCode(pin_code, pin_code_len)
        # if addr == 0:
        #     print('连接设备的蓝牙地址错误！')
        #     return 0
        # # print('搜索蓝牙地址为：',end='')
        # # print(addr)
        #
        # bdaddr = (c_char * 6)(*addr)
        # Search_flag = self.sdk_dll.BluetSearchRemoteDevice(brclass, bdaddr)
        # print('模块搜索中！')
        # while True:
        #     if self.search_flag == 1:
        #         self.search_flag = 0
        #         print('已经搜索到设备')
        #         self.sdk_dll.BluetStopDiscoverDevice()
        #         print('停止搜索')
        #         break
        # print(self.dev_hdl)
        # self.sdk_dll.GATT_Connect.argtypes = [c_ulong]
        #
        # flag = self.sdk_dll.GATT_Connect(self.dev_hdl)
        # print(self.dev_hdl)
        # print('连接标志：',end='')
        # print(flag)

        flag = self.Gatt_Connection(bd2)
        if flag == 1:
            print('连接成功')
        else:
            print('连接失败')
        self.sdk_dll.GATT_GetServices.restype = POINTER(BtSdkGATTService)
        gatt_service = self.sdk_dll.GATT_GetServices()                          # 得到服务 get services
        self.sdk_dll.ATT_GETUUIDType.argtypes = [c_ushort]
        self.sdk_dll.ATT_GETUUIDType.restype = c_char_p

        self.sdk_dll.GATT_GetCharacteristics.argtypes = [c_ubyte]               # 得到特征 get characteristics
        self.sdk_dll.GATT_GetCharacteristics.restype = POINTER(BtSdkGATTCharacteristic)

        self.sdk_dll.GATT_GetCharacteristicsValue.argtypes = [c_ubyte]          # 得到特征值 get characteristics_value
        self.sdk_dll.GATT_GetCharacteristicsValue.restype = POINTER(BtsdsGATTCharacteristicValueStru)

        self.sdk_dll.GATT_SetCharacteristicsValue.argtypes = [c_char, c_char]    # 设置特征值 set characteristics_value
        self.sdk_dll.GATT_SetCharacteristicsValue.retypes = c_long

        self.sdk_dll.GATT_GetDescriptors.argtypes = [c_ubyte]                   # 得到描述 get Descriptors
        self.sdk_dll.GATT_GetDescriptors.restype = POINTER(BtSdkGATTDescriptor)

        self.sdk_dll.GATT_GetDescriptorValue.argtypes = [c_ubyte]               # 得到描述值 get Descriptors_value
        self.sdk_dll.GATT_GetDescriptorValue.restype = POINTER(BtsdkGATTDscriptorValueStru)

        self.sdk_dll.GATT_SetDescriptorValue.argtypes = [c_ubyte, c_ubyte]      # 设置描述值 set Descriptors_value

        try:
            if gatt_service is None:
                print('没有发现服务')
                return 0
            else:
                # print('gatt_service')
                # print(gatt_service.contents.num)       # 服务个数
                for i in range(int(gatt_service.contents.num)):
                    uuid = gatt_service.contents.service[i].ServiceUuid.ShortUuid
                    src = self.sdk_dll.ATT_GETUUIDType(uuid)
                    print("service id:%d  uuid = %d %s handle = 0x%02x" % (
                        i+1, uuid, src.decode(), gatt_service.contents.service[i].AttributeHandle))

                    gatt_character = self.sdk_dll.GATT_GetCharacteristics(c_ubyte(i+1))
                    if gatt_character is not None:
                        # print('     gatt_character')
                        # print("gatt_character: " + str(gatt_character.contents.num))
                        for j in range(gatt_character.contents.num):
                            uuid_c = gatt_character.contents.character[j].CharacteristicUuid.ShortUuid
                            src_c = self.sdk_dll.ATT_GETUUIDType(uuid_c)

                            print("     character id:%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
                                j + 1, gatt_character.contents.character[j].AttributeHandle,
                                gatt_character.contents.character[j].CharacteristicValueHandle, src_c.decode()))
                            print("     [B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
                                gatt_character.contents.character[j].IsBroadcastable,
                                gatt_character.contents.character[j].IsReadable,
                                gatt_character.contents.character[j].IsWritable,
                                gatt_character.contents.character[j].IsSignedWritable,
                                gatt_character.contents.character[j].IsWritableWithoutRespon,
                                gatt_character.contents.character[j].IsIndicatable,
                                gatt_character.contents.character[j].IsNotifiable,
                                gatt_character.contents.character[j].HasExtendedProperties))

                            gatt_character_val = self.sdk_dll.GATT_GetCharacteristicsValue(c_ubyte(j+1))
                            if gatt_character_val is not None and gatt_character.contents.character[j].IsReadable == 1:
                                # print('     gatt_character_val')
                                for s in range(gatt_character_val.contents.DataSize):
                                    print("     value:0x%02x" % gatt_character_val.contents.Data[s],end='')
                                print('')

                            gatt_des = self.sdk_dll.GATT_GetDescriptors(c_ubyte(j+1))
                            try:
                                if gatt_des is not None:
                                    # print('             Descriptor:')
                                    for n in range(gatt_des.contents.num):
                                        uuid = gatt_des.contents.des[i].DescriptorUuid.ShortUuid
                                        src = self.sdk_dll.ATT_GETUUIDType(uuid)
                                        print("             Descriptor id:%d,[Hdl,D:0x%04x] %s" % (n,gatt_des.contents.des[i].AttributeHandle, src.decode()))
                                        print("             character handle ="+gatt_des.contents.des[i].CharacteristicHandle)

                                        gatt_des_val = self.sdk_dll.GATT_GetDescriptorValue(c_ubyte(n+1))
                                        if gatt_des_val is not None:
                                            # print('             gatt_des_val')
                                            if gatt_des_val.contents.DescriptorType == 2:
                                                print("             Notify:%d,Ind:%d" % (
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToNotification,
                                                    gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToIndication))
                            except Exception as e:
                                print('',end='')
                                # print('没有Descriptors')
        except Exception as e:
            print(e)
            print('发生错误')
            return 0

        # sel_service = c_ubyte(1)
        # gatt_character = self.sdk_dll.GATT_GetCharacteristics(sel_service)
        # if gatt_character is None:
        #     pass
        # try:
        #     print("gatt_character: " + str(gatt_character.contents.num))
        #     for j in range(gatt_character.contents.num):
        #         uuid = gatt_character.contents.character[j].CharacteristicUuid.ShortUuid
        #         src = self.sdk_dll.ATT_GETUUIDType(uuid)
        #         print("character id :%d,[Hdl,D:0x%08x,V:0x%08x] %s" % (
        #             j + 1, gatt_character.contents.character[j].AttributeHandle,
        #             gatt_character.contents.character[j].CharacteristicValueHandle, src.decode()))
        #         print("[B:%d R:%d W:%d SW:%d WWR:%d I:%d N:%d E:%d]" % (
        #             gatt_character.contents.character[j].IsBroadcastable,
        #             gatt_character.contents.character[j].IsReadable,
        #             gatt_character.contents.character[j].IsWritable,
        #             gatt_character.contents.character[j].IsSignedWritable,
        #             gatt_character.contents.character[j].IsWritableWithoutRespon,
        #             gatt_character.contents.character[j].IsIndicatable,
        #             gatt_character.contents.character[j].IsNotifiable,
        #             gatt_character.contents.character[j].HasExtendedProperties))
        # except Exception as e:
        #     print(e)
        # self.sdk_dll.GATT_GetCharacteristicsValue.argtypes = [c_ubyte]
        # self.sdk_dll.GATT_GetCharacteristicsValue.restype = POINTER(BtsdsGATTCharacteristicValueStru)
        # sel_character = c_ubyte(1)
        # character_val = self.sdk_dll.GATT_GetCharacteristicsValue(sel_character)
        #
        # try:
        #     for i in range(character_val.contents.DataSize):
        #         print("value:0x%02x" % character_val.contents.Data[i])
        # except Exception as e:
        #     print(e)
        # self.sdk_dll.GATT_SetCharacteristicsValue.argtypes = [c_ubyte, c_char_p]
        # self.sdk_dll.GATT_SetCharacteristicsValue.restype = c_char
        # sel_crt = c_ubyte(1)
        # in_str = c_char_p(bytes("01", 'utf-8'))
        # self.sdk_dll.GATT_SetCharacteristicsValue(sel_crt, in_str)
        # self.sdk_dll.GATT_GetDescriptors.argtypes = [c_ubyte]
        # self.sdk_dll.GATT_GetDescriptors.restype = POINTER(BtSdkGATTDescriptor)
        # sel_crt = c_ubyte(1)
        # gatt_des = self.sdk_dll.GATT_GetDescriptors(sel_crt)
        # for i in range(gatt_des.contents.num):
        #     uuid = gatt_des.contents.des[i].DescriptorUuid.ShortUuid
        #     src = self.sdk_dll.ATT_GETUUIDType(uuid)
        #     print(" [Hdl,D:0x%04x] %s" % (gatt_des.contents.des[i].AttributeHandle, src.decode()))
        #     print("character handle = ", gatt_des.contents.des[i].CharacteristicHandle)
        # self.sdk_dll.GATT_GetDescriptorValue.argtypes = [c_ubyte]
        # self.sdk_dll.GATT_GetDescriptorValue.restype = POINTER(BtsdkGATTDscriptorValueStru)
        # sel_des = c_ubyte(1)
        # gatt_des_val = self.sdk_dll.GATT_GetDescriptorValue(sel_des)
        #
        # if gatt_des_val.contents.DescriptorType == 2:
        #     print("Notify:%d,Ind:%d" % (
        #         gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToNotification,
        #         gatt_des_val.contents.u.ClientCharacteristicConfiguration.IsSubscribeToIndication))
        # self.sdk_dll.GATT_SetDescriptorValue.argtypes = [c_ubyte, c_ubyte]
        # sel_dec = c_ubyte(1)
        # val = c_ubyte(1)
        # self.sdk_dll.GATT_SetDescriptorValue(sel_dec, val)
        # self.sdk_dll.GATT_AutoCache.retypes = c_char
        # s = self.sdk_dll.GATT_AutoCache(self.dev_hdl)  # 将gatt服务写入到txt文本当中
        # print('写入标志：', end='')
        # print(s)

        # self.sdk_dll.GATT_DeInit()
        # self.sdk_dll.GATT_EndSession(self.dev_hdl)

    def test_excel(self, result, question):
        read_data = openpyxl.load_workbook(self.path)
        # print(read_data.get_sheet_names())  # 输出所有工作页的名称
        table = read_data.get_sheet_by_name(self.sheet_name)
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        for i in range(1, ncolumns+1):
            if table.cell(1, i).value == self.result:
                for j in range(2, nrows+1):
                    # print(result[j-2])
                    if result[j-2] != ' ':
                        table.cell(j, i).value = str(result[j-2])
                        # print(table.cell(j, i).value)
            if table.cell(1, i).value == self.question:
                for j in range(2, nrows+1):
                    # print(question[j-2])
                    if question[j-2] != ' ':
                        table.cell(j, i).value = str(question[j-2])
                        # print(table.cell(j, i).value)
        try:
            read_data.save('Test_Report.xlsx')
        except:
            self.log.write('报告生成文件未关闭！', self.scr)

    def creat_report(self, index1, result, index2, question):
        '''
        生成测试报告
        :return:
        '''
        if not os.path.exists(self.report_path):
            f = open(self.report_path, 'w+')
            f.close()
        data = pd.read_excel(self.path, sheet_name='SPP')
        print(data)
        # for i in self.type():
        #     print(data[i])
        # for i in range(len(data[self.type[0]])):
        #     data.loc[i, self.type[index1]] = result[i]
        #     data.loc[i, self.type[index2]] = question[i]

        # print('asdf')
        # data = data.loc[:, ~data.columns.str.contains('Unnamed')]  # 去除Unnamed列
        # for i in self.type():
        #     print(data[i])
        # data.to_excel(self.report_path, sheet_name='Report', index=False)


if __name__ == "__main__":
    tool = Auto_Test()
    # sys.stdout = Log()
    # tool.test_log()
    tool.creat_map()
    # tool.del_data()
    # tool.creat_report()
    # tool.test_gatt()
    # tool.spp_t()
    # tool.gatt_t()
    # tool.test_spp()